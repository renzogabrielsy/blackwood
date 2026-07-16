#!/usr/bin/env python3
"""
FLECON bag-movement extractor — the "Bagging Manager" (v1). See FLECON_BAGGING_DESIGN.md.

This ingests packaging-material (empty jumbo/flecon bag) stock — NOT charcoal. Source is
Ivy's single CUMULATIVE workbook `FLECON BAG MOVEMENT 2026.xlsx`, ONE tab per YEAR
(`JANUARY 2026` = all of 2026, with in-sheet month section headers). It is a sibling of
extract_daily_production.py (same coercion helpers, same argparse/--since/JSON shape,
same date carry-forward idea). Deterministic muscle only — the sheet parse itself is
offline (NO network), but bag-type COLUMN MAPPING is driven by the DB registry
(`flecon_bag_types`) fetched read-only via lib/db.py so the column map survives the
operator reshuffling / renaming columns. JSON to stdout (+ optional --work-dir file).

RESILIENCE MODEL (2026-07-02 rewrite) — HEADER-SIGNATURE MATCHING, position-independent:
Previously each bag type was pinned to a FIXED column letter (C->P). That silently broke
if the operator inserted/moved a column. Now each data column is matched by its HEADER
SIGNATURE (the combined header text) against the registry's `source_label`. New /
unrecognized columns are FLAGGED for the user to register — never auto-created, never
silently dropped or misassigned (same posture as unmapped batches elsewhere in this skill).

Workbook structure (verified 2026-07-02 against tab "JANUARY 2026", ~510 rows):
- Header is MULTI-ROW. Row 4 A='DATE', B='PARTICULAR', C4='FLECON BAG' (block label),
  H4='Running Balance' (a label, not data). Bag-type identity is spread across header
  rows 3 / 5 / 6: e.g. col M = row3 '6X50' + row5 'FG' + row6 'w/ Black Sling';
  col L = row5 'ECOPACK BEIGE' + row6 'TUNNER BAG'. A single header cell is ambiguous
  (M and N both say 'FG'; K and L both say 'ECOPACK BEIGE'), so we COMBINE header rows
  3/5/6 for a column into one signature to disambiguate.
- Row 7 'Forwarded Balance': per-bag-type OPENING stock for the year (blanks = 0), now
  keyed by the SAME signature-matched columns (not fixed letters).
- Data rows: A=date (a dated cell starts a date; sub-rows inherit it — carry-forward),
  B=particular (event text), and a SIGNED integer in one bag-type column (99.6% single
  column; rare rows move two — a blend/recount). Negative = bags consumed OUT; positive
  = bags received/returned IN. Emit ONE movement per populated MATCHED bag-type column.
- Month section rows: A=month name (alpha), B empty -> context marker, skip (also RESET the
  carried date so a month header can't leak the prior month's date into a bare sub-row).
- Marker rows: many `RS 1 ZAMBOANGA`/`RS 1 ZAMBAONGA` rows carry NO bag quantity -> skip
  (not movements). Keep both raw spellings verbatim; do NOT auto-"fix".
- Balance snapshot row (~row 499): per-type running balances the operator maintains — NO
  date and NO particular, but numbers across many columns. It is NOT a movement: the
  "particular present" guard skips it, and it is captured separately as balance_snapshot
  for the informational cross-check.

Usage:
    python3 extract_flecon_bags.py --file "FLECON BAG MOVEMENT 2026.xlsx" --since 2026-01-01
    python3 extract_flecon_bags.py --file "..." --since 2026-06-29        # tail-scope
    python3 extract_flecon_bags.py --file "..." --year 2026               # force year tab
    python3 extract_flecon_bags.py --file "..." --since 2026-01-01 --work-dir /tmp/run
    # Offline (no DB): supply the registry as JSON instead of hitting Supabase.
    python3 extract_flecon_bags.py --file "..." --since ... --bag-types-json /tmp/types.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)

# lib/db.py lives alongside this script under lib/ — used ONLY to fetch the bag-type
# registry (read-only). Offline mode via --bag-types-json still works if it's unavailable.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from lib.db import DBClient  # type: ignore
except Exception:  # noqa: BLE001
    DBClient = None  # type: ignore


# ---------------------------------------------------------------------------
# Header geometry (design §2). The bag-type block starts at column C. Bag-type identity
# is combined from header rows 3, 5, 6 into ONE signature per column (row order is stable
# so signatures are deterministic). We scan the bag-type block bounded to the registry's
# authored column span (C .. max source_column, default C..P) — matching is by SIGNATURE
# within that block, so a reshuffle of the SKUs is handled; a column in-block that matches
# no registry entry but carries qty is FLAGGED (candidate new bag type), never guessed.
# ---------------------------------------------------------------------------
COL_C = 3                    # 'C' — first bag-type column
DEFAULT_LAST_BAG_COL = 16    # 'P' — historical right edge of the 14-SKU block
HEADER_SIGNATURE_ROWS = (3, 5, 6)   # combined, in this order, into a column's signature

COL_DATE = 1        # A
COL_PARTICULAR = 2  # B
OPENING_BALANCE_ROW = 7    # "Forwarded Balance"
FIRST_DATA_ROW = 8

MONTH_NAMES = {
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
}


# ---------------------------------------------------------------------------
# Registry (flecon_bag_types) — the source of truth for column mapping.
# ---------------------------------------------------------------------------
def load_bag_type_registry(bag_types_json: str | None) -> list[dict[str, Any]]:
    """
    Fetch the bag-type registry: [{code, source_label, source_column, sort_order, label}].
    From --bag-types-json when given (offline), else read-only via lib/db.py. Raises with a
    clear message if neither path yields the registry (the column map depends on it).
    """
    if bag_types_json:
        rows = json.loads(Path(bag_types_json).expanduser().read_text())
        if isinstance(rows, list) and len(rows) == 1 and isinstance(rows[0], dict) and "data" in rows[0]:
            rows = rows[0]["data"] or []
        return rows
    if DBClient is None:
        raise RuntimeError(
            "flecon_bag_types registry unavailable: lib.db could not be imported and no "
            "--bag-types-json was supplied. Cannot build the column map."
        )
    return DBClient().read_rows(
        "flecon_bag_types",
        columns=["code", "source_label", "source_column", "sort_order", "label"],
        since_date=None,
    )


def _col_letter_to_index(letter: Any) -> int | None:
    """'C' -> 3. Tolerant: strips spaces, uppercases; None/blank -> None."""
    if letter is None:
        return None
    s = str(letter).strip().upper()
    if not s or not s.isalpha():
        return None
    idx = 0
    for ch in s:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def normalize_sig(text: str | None) -> str:
    """
    Canonical signature key: lowercase, then DROP every non-alphanumeric char (spaces,
    punctuation, parentheses, slashes, dashes). '590 kls (Kuraray)' -> '590klskuraray'.
    Used to compare a column's combined header signature against a registry source_label.
    """
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


# ---------------------------------------------------------------------------
# Coercion helpers (mirrors extract_daily_production.py)
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> int | None:
    f = coerce_float(value)
    return int(round(f)) if f is not None else None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_since(value: str | None) -> date | None:
    """Parse optional --since 'YYYY-MM-DD'. None when omitted. Raises ValueError if malformed."""
    if value is None:
        return None
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


# ---------------------------------------------------------------------------
# Sheet selection — pick the tab for the target year.
# ---------------------------------------------------------------------------
def select_year_sheet(wb, year: int | None) -> tuple[str, str | None]:
    """
    Pick the sheet whose name contains the target year (e.g. 'JANUARY 2026' or 'JANUARY2026').
    Fallback: the LAST sheet in the workbook (the cumulative current-year tab is authored last).
    Returns (sheet_name, warning-or-None). NOTE: cell E2 may carry a stale 'YEAR 2025' label —
    we trust the tab name + row dates, never E2 (design §1).
    """
    if year is not None:
        target = str(year)
        matches = [n for n in wb.sheetnames if target in n.replace(" ", "")]
        if matches:
            # Prefer the last match (most recent authored) if several contain the year.
            return matches[-1], None
        return wb.sheetnames[-1], (
            f"No sheet name contains year {year}; falling back to last sheet "
            f"'{wb.sheetnames[-1]}'. Available: {wb.sheetnames}"
        )
    # No year given: use the last sheet (the cumulative current-year tab).
    return wb.sheetnames[-1], None


# ---------------------------------------------------------------------------
# Header block location (hard-error if it can't be found)
# ---------------------------------------------------------------------------
def locate_header_block(ws) -> bool:
    """
    Verify the DATE/PARTICULAR header block is where we expect (row 4 A='DATE',
    B='PARTICULAR') AND that at least one bag-type-area column carries a header signature.
    Returns True if located. Callers hard-error (don't silently produce 0 rows) on False.
    """
    a4 = normalize_sig(coerce_str(ws.cell(4, COL_DATE).value))
    b4 = normalize_sig(coerce_str(ws.cell(4, COL_PARTICULAR).value))
    header_ok = a4 == "date" and b4 == "particular"
    any_sig = any(
        build_column_signature(ws, c).strip()
        for c in range(COL_C, min(DEFAULT_LAST_BAG_COL, ws.max_column) + 1)
    )
    return header_ok and any_sig


# ---------------------------------------------------------------------------
# Column mapping by HEADER SIGNATURE (the resilience core)
# ---------------------------------------------------------------------------
def build_column_signature(ws, col: int) -> str:
    """Combine this column's non-empty header cells across HEADER_SIGNATURE_ROWS (3/5/6),
    in that stable order, into one space-joined signature string."""
    parts: list[str] = []
    for r in HEADER_SIGNATURE_ROWS:
        v = coerce_str(ws.cell(r, col).value)
        if v:
            parts.append(v)
    return " ".join(parts)


def map_columns(
    ws, registry: list[dict[str, Any]], warnings: list[str]
) -> tuple[dict[int, str], list[dict[str, Any]], list[str]]:
    """
    Match each bag-type-area column (C .. last authored bag column) to a registry `code`
    by comparing its combined header signature to each entry's normalized `source_label`.

    Two passes, conservative:
      1. EXACT normalized match — signature == source_label (after normalize_sig).
      2. CONTAINS fallback for still-unmatched columns — a registry label fully contained
         in the column's signature (or vice-versa), but ONLY when it is UNAMBIGUOUS: the
         candidate code must not already be claimed by an exact match, and exactly one
         registry entry may contain-match. Any ambiguity ⇒ leave the column UNMAPPED.

    Returns:
      col_to_code:  {col_index: code} for matched columns.
      column_map:   [{column_letter, signature, matched_code|null, sort_order}] for every
                    bag-type-area column found (matched or not).
      matched_codes: list of codes that got matched (for missing_columns computation).
    """
    # Build normalized registry lookup. source_column is only a tiebreak hint.
    reg_by_nsig: dict[str, list[dict[str, Any]]] = {}
    for e in registry:
        nsig = normalize_sig(e.get("source_label"))
        if nsig:
            reg_by_nsig.setdefault(nsig, []).append(e)
    reg_entries = [e for e in registry if normalize_sig(e.get("source_label"))]

    # Determine the right edge of the bag-type block: max registry source_column, but at
    # least the historical 'P'. Scan C .. that edge.
    reg_cols = [
        idx for idx in (_col_letter_to_index(e.get("source_column")) for e in registry)
        if idx is not None
    ]
    last_col = max([DEFAULT_LAST_BAG_COL, *reg_cols]) if reg_cols else DEFAULT_LAST_BAG_COL

    col_to_code: dict[int, str] = {}
    claimed_codes: set[str] = set()
    column_map: list[dict[str, Any]] = []
    # First hold per-column signature + normalized form so we can do pass 2.
    scanned: list[tuple[int, str, str]] = []  # (col, signature, nsig)

    for col in range(COL_C, last_col + 1):
        sig = build_column_signature(ws, col)
        nsig = normalize_sig(sig)
        scanned.append((col, sig, nsig))

    # Pass 1 — exact normalized match.
    for col, sig, nsig in scanned:
        if not nsig:
            continue
        hits = reg_by_nsig.get(nsig, [])
        if len(hits) == 1:
            code = str(hits[0]["code"])
            col_to_code[col] = code
            claimed_codes.add(code)
        elif len(hits) > 1:
            warnings.append(
                f"col {get_column_letter(col)}: signature {sig!r} matched {len(hits)} "
                f"registry entries exactly ({[h['code'] for h in hits]}) — left UNMAPPED (ambiguous)."
            )

    # Pass 2 — conservative contains fallback for still-unmatched columns.
    for col, sig, nsig in scanned:
        if col in col_to_code or not nsig:
            continue
        cand: list[dict[str, Any]] = []
        for e in reg_entries:
            code = str(e["code"])
            if code in claimed_codes:
                continue  # already taken by an exact match elsewhere
            rn = normalize_sig(e.get("source_label"))
            if rn and (rn in nsig or nsig in rn):
                cand.append(e)
        # Deduplicate candidate codes.
        cand_codes = sorted({str(e["code"]) for e in cand})
        if len(cand_codes) == 1:
            code = cand_codes[0]
            col_to_code[col] = code
            claimed_codes.add(code)
        elif len(cand_codes) > 1:
            warnings.append(
                f"col {get_column_letter(col)}: signature {sig!r} contains-matched "
                f"{cand_codes} — left UNMAPPED (ambiguous, never guessed)."
            )

    # Build the column_map report entry for every scanned bag-type-area column.
    sort_by_code = {str(e["code"]): e.get("sort_order") for e in registry}
    for col, sig, nsig in scanned:
        code = col_to_code.get(col)
        column_map.append({
            "column_letter": get_column_letter(col),
            "signature": sig,
            "matched_code": code,
            "sort_order": sort_by_code.get(code) if code else None,
        })

    matched_codes = list(col_to_code.values())
    return col_to_code, column_map, matched_codes


def scan_unmapped_columns(
    ws, col_to_code: dict[int, str], column_map: list[dict[str, Any]],
    since: date | None,
) -> list[dict[str, Any]]:
    """
    Any bag-type-area column that matched NO registry entry BUT carries >=1 non-empty qty
    cell in the scanned range is a CANDIDATE NEW bag type — surface it (never emit as a
    movement, never guess a code). Returns [{column_letter, signature, sample_values,
    first_data_row}].
    """
    letter_to_col = {get_column_letter(c): c for c in range(COL_C, ws.max_column + 1)}
    out: list[dict[str, Any]] = []
    for entry in column_map:
        if entry["matched_code"] is not None:
            continue
        letter = entry["column_letter"]
        col = letter_to_col.get(letter)
        if col is None:
            continue
        samples: list[dict[str, Any]] = []
        first_row: int | None = None
        # Include the opening-balance row (7) and data rows.
        opening = coerce_int(ws.cell(OPENING_BALANCE_ROW, col).value)
        if opening is not None and opening != 0:
            first_row = OPENING_BALANCE_ROW
            samples.append({"row": OPENING_BALANCE_ROW, "value": opening, "kind": "opening"})
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            iv = coerce_int(ws.cell(r, col).value)
            if iv is not None and iv != 0:
                if first_row is None:
                    first_row = r
                if len(samples) < 5:
                    samples.append({"row": r, "value": iv})
                else:
                    break
        if samples:
            out.append({
                "column_letter": letter,
                "signature": entry["signature"],
                "sample_values": samples,
                "first_data_row": first_row,
            })
    return out


# ---------------------------------------------------------------------------
# Opening balances (row 7) — keyed by the SIGNATURE-matched columns
# ---------------------------------------------------------------------------
def extract_opening_balances(ws, col_to_code: dict[int, str]) -> dict[str, int]:
    """Parse row 7 'Forwarded Balance' -> {code: qty} for MATCHED columns. Blank = 0 (omitted)."""
    out: dict[str, int] = {}
    for col, code in col_to_code.items():
        v = coerce_int(ws.cell(OPENING_BALANCE_ROW, col).value)
        if v is not None and v != 0:
            out[code] = v
    return out


# ---------------------------------------------------------------------------
# Movement extraction
# ---------------------------------------------------------------------------
def extract_movements(
    ws,
    col_to_code: dict[int, str],
    since: date | None,
    warnings: list[str],
) -> tuple[list[dict[str, Any]], dict[str, int] | None, int, int]:
    """
    Walk the data rows carrying the date forward, emitting one movement per populated
    MATCHED bag-type column (the bag_type_code now comes from signature matching, not a
    fixed letter). Unmatched columns are never emitted here (they are surfaced separately
    as unmapped_columns). Returns
    (movements, balance_snapshot_or_None, dropped_before_since, skipped_markers).
    """
    movements: list[dict[str, Any]] = []
    balance_snapshot: dict[str, int] | None = None
    carried_date: date | None = None
    dropped_before_since = 0
    skipped_markers = 0
    matched_cols = sorted(col_to_code)

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        a_raw = ws.cell(r, COL_DATE).value
        particular = coerce_str(ws.cell(r, COL_PARTICULAR).value)

        # Populated MATCHED bag-type columns for this row (signed ints).
        cols: list[tuple[int, int]] = []
        for c in matched_cols:
            iv = coerce_int(ws.cell(r, c).value)
            if iv is not None and iv != 0:
                cols.append((c, iv))

        a_str = coerce_str(a_raw)
        a_date = coerce_date(a_raw)

        # Month-name section row: col A alpha (month), col B empty -> context marker.
        # Reset carried date so a month header can't leak the prior month's date forward.
        if a_str is not None and a_date is None and particular is None:
            if a_str.upper() in MONTH_NAMES:
                carried_date = None
                continue
            # Any other alpha-in-A / no-particular row (e.g. the 'NOTE:' footer) is not a movement.
            # If it also has no quantities, skip silently; if it has quantities, treat as snapshot below.

        # Balance-snapshot row: NO date, NO particular, but numbers present (~row 499).
        # Capture once for the informational cross-check; never emit as movements.
        if a_date is None and particular is None and cols:
            if balance_snapshot is None:
                balance_snapshot = {}
                for c, v in cols:
                    balance_snapshot[col_to_code[c]] = v
            else:
                warnings.append(f"row {r}: second balance-snapshot-like row ignored")
            continue

        # Carry the date forward: a dated cell starts a new date; sub-rows inherit it.
        if a_date is not None:
            carried_date = a_date

        # No matched bag-type quantity on this row -> a bare marker (e.g. RS 1 ZAMBOANGA). Skip.
        if not cols:
            if particular is not None:
                skipped_markers += 1
            continue

        # A movement needs an effective date. If none carried yet, warn and skip.
        if carried_date is None:
            warnings.append(
                f"row {r}: bag quantity present but no date in context "
                f"(particular={particular!r}) — skipped"
            )
            continue

        # --since tail-scope: drop rows dated before the watermark.
        if since is not None and carried_date < since:
            dropped_before_since += len(cols)
            continue

        # Emit one movement per populated matched column (handles the rare multi-column row).
        for c, qty in cols:
            movements.append({
                "transaction_date": carried_date.isoformat(),
                "particular": particular,          # raw verbatim (both ZAMBOANGA spellings kept)
                "bag_type_code": col_to_code[c],
                "qty_delta": qty,                  # signed int: - = OUT, + = IN
                "source_row": r,
            })

    return movements, balance_snapshot, dropped_before_since, skipped_markers


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract FLECON bag movements from Ivy's cumulative bag-movement workbook."
    )
    parser.add_argument("--file", required=True, help="Path to FLECON BAG MOVEMENT <year>.xlsx")
    parser.add_argument("--since", default=None,
                        help="Watermark 'YYYY-MM-DD'. Drop movements dated STRICTLY BEFORE this "
                             "date. First run = full-year backfill uses --since 2026-01-01; daily "
                             "runs tail-scope with --since = watermark - 3 days. Omit for no filter.")
    parser.add_argument("--year", type=int, default=None,
                        help="Force the year tab (sheet name containing this year). "
                             "Default: last sheet in the workbook.")
    parser.add_argument("--work-dir", default=None,
                        help="Optional dir to also write the JSON payload (extract_flecon_bags.json).")
    parser.add_argument("--bag-types-json", default=None,
                        help="Offline registry override: JSON list of flecon_bag_types rows "
                             "[{code, source_label, source_column, sort_order, label}]. "
                             "Default: fetched read-only via lib/db.py.")
    args = parser.parse_args()

    try:
        since = parse_since(args.since)
    except ValueError:
        print(json.dumps({
            "error": f"Invalid --since '{args.since}'. Expected format YYYY-MM-DD."
        }), file=sys.stderr)
        return 2

    path = Path(args.file).expanduser()
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    warnings: list[str] = []
    sheet_name, sel_warn = select_year_sheet(wb, args.year)
    if sel_warn:
        warnings.append(sel_warn)
    ws = wb[sheet_name]

    # Hard-error if the header block itself can't be located (never silently emit 0 rows).
    if not locate_header_block(ws):
        print(json.dumps({
            "error": (
                f"Could not locate the FLECON header block on sheet '{sheet_name}' "
                f"(expected row 4 A='DATE'/B='PARTICULAR' and bag-type header signatures "
                f"from row {COL_C}). Aborting rather than producing 0 movements."
            )
        }), file=sys.stderr)
        return 4

    # Load the bag-type registry (source of truth for the column map).
    try:
        registry = load_bag_type_registry(args.bag_types_json)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"error": f"Failed to load flecon_bag_types registry: {e}"}),
              file=sys.stderr)
        return 5
    if not registry:
        print(json.dumps({"error": "flecon_bag_types registry is empty — cannot map columns."}),
              file=sys.stderr)
        return 5

    # Signature-match each bag-type-area column to a registry code (position-independent).
    col_to_code, column_map, matched_codes = map_columns(ws, registry, warnings)
    unmapped_columns = scan_unmapped_columns(ws, col_to_code, column_map, since)

    # missing_columns: registry codes whose source_label matched NO column this run (warn).
    matched_set = set(matched_codes)
    missing_columns = [
        {"code": str(e["code"]), "source_label": e.get("source_label"),
         "source_column": e.get("source_column")}
        for e in registry if str(e["code"]) not in matched_set
    ]
    if unmapped_columns:
        warnings.append(
            f"{len(unmapped_columns)} unmapped column(s) with data — possible NEW bag type(s), "
            f"FLAGGED for registration: {[u['column_letter'] for u in unmapped_columns]}"
        )
    if missing_columns:
        warnings.append(
            f"{len(missing_columns)} registry code(s) matched NO column this run "
            f"(removed/renamed?): {[m['code'] for m in missing_columns]}"
        )

    opening_balances = extract_opening_balances(ws, col_to_code)
    movements, balance_snapshot, dropped_before_since, skipped_markers = extract_movements(
        ws, col_to_code, since, warnings
    )

    # Confidence: 1.0 minus a small penalty per warning, floored at 0.5.
    overall_confidence = round(max(0.5, 1.0 - 0.05 * len(warnings)), 3)

    # Distinct dates + signed sanity totals for the summary.
    dates = sorted({m["transaction_date"] for m in movements})
    total_in = sum(m["qty_delta"] for m in movements if m["qty_delta"] > 0)
    total_out = sum(-m["qty_delta"] for m in movements if m["qty_delta"] < 0)

    output = {
        "filename": path.name,
        "sheet": sheet_name,
        "since": args.since,
        "rows": movements,
        "opening_balances": opening_balances,
        "balance_snapshot": balance_snapshot,   # sheet's own running balance (informational)
        "column_map": column_map,               # every bag-type-area column + what it matched
        "unmapped_columns": unmapped_columns,    # candidate NEW bag types → FLAG, never emit
        "missing_columns": missing_columns,      # registry codes with no column this run → warn
        "summary": {
            "total_rows": len(movements),
            "distinct_dates": len(dates),
            "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None,
            "total_in": total_in,
            "total_out": total_out,
            "dropped_before_since": dropped_before_since,
            "skipped_markers": skipped_markers,
            "matched_columns": len(col_to_code),
            "unmapped_columns": len(unmapped_columns),
            "missing_columns": len(missing_columns),
            "extraction_warnings": warnings,
            "overall_confidence": overall_confidence,
        },
    }

    payload = json.dumps(output, indent=2, default=str)
    if args.work_dir:
        wd = Path(args.work_dir).expanduser()
        wd.mkdir(parents=True, exist_ok=True)
        (wd / "extract_flecon_bags.json").write_text(payload)
    print(payload)
    return 0


if __name__ == "__main__":
    sys.exit(main())
