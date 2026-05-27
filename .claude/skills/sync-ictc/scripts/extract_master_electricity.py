#!/usr/bin/env python3
"""
MASTER ELECTRICITY sheet extractor — backfill script for electricity_readings table.

Source: MASTER - ICTC INPUT FILE V1.xlsx, sheet "ELECTRICITY"
Structure: Two sections, both are extracted:

  SECTION A — Monthly summaries (rows 3-14): 2026 first-of-month aggregates.
    Col 1=MONTH, Col 3=TOTAL START, Col 4=TOTAL END, Col 5=TOTAL DIFF, Col 6=RATE, Col 7=TTL KHW
    Col 10=BUNKHOUSE START, Col 11=BUNKHOUSE END, Col 12=BUNKHOUSE DIFF
    Col 17=PUMP START, Col 18=PUMP END, Col 19=PUMP DIFF
    Skip rows where DIFF == 0 (future months, not yet active).

  SECTION B — Daily detail (rows 19+): 2025+ daily readings.
    Same column layout but one row per day. Headers at rows 16-18.
    Col 1=PROD DAY, Col 3=PLANT START, Col 4=PLANT END, Col 5=PLANT DIFF
    Col 10=BUNKHOUSE START, Col 11=BUNKHOUSE END, Col 12=BUNKHOUSE DIFF
    Col 17=PUMP START, Col 18=PUMP END, Col 19=PUMP DIFF
    Skip rows where start AND end are both null/0 (no reading recorded).

Both sections feed into electricity_readings with the meter name as:
  TOTAL/MAIN -> meter='MAIN'
  BUNKHOUSE  -> meter='BUNKHOUSE'
  PUMP       -> meter='PUMP'

DB schema: UNIQUE(reading_date, meter). Monthly 2026 rows get reading_date = first-of-month.
Daily 2025 rows get reading_date = that calendar date.

Usage:
    python3 extract_master_electricity.py --file "path/to/MASTER.xlsx"
    python3 extract_master_electricity.py --file "..." --output /tmp/electricity_extract.json

Output:
    {
      "source_file": "...",
      "source_sheet": "ELECTRICITY",
      "rows": [ {electricity_readings row}, ... ],
      "summary": {
        "total_rows": N,
        "monthly_rows": N,
        "daily_rows": N,
        "date_range": "YYYY-MM-DD -> YYYY-MM-DD",
        "meters": {"MAIN": N, "BUNKHOUSE": N, "PUMP": N},
        "warnings": [...],
      }
    }
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)


# ---------------------------------------------------------------------------
# Column indices (1-based)
# ---------------------------------------------------------------------------
COL_DATE         = 1

# TOTAL / MAIN meter (cols 3-7 in section A, cols 3-8 in section B)
COL_TOTAL_START  = 3
COL_TOTAL_END    = 4
COL_TOTAL_DIFF   = 5
COL_RATE         = 6   # Always 120 PHP/kWh for MAIN; section B row 17 has it hardcoded

# BUNKHOUSE meter (cols 10-14 in section A, cols 10-15 in section B)
COL_BH_START     = 10
COL_BH_END       = 11
COL_BH_DIFF      = 12

# PUMP meter (cols 17-21 in section A, cols 17-22 in section B)
COL_PUMP_START   = 17
COL_PUMP_END     = 18
COL_PUMP_DIFF    = 19

# Section boundaries
MONTHLY_DATA_START = 3
MONTHLY_DATA_END   = 14   # rows 3-14 = JAN-DEC 2026 monthly summaries
DAILY_DATA_START   = 19   # row 19 onward = daily detail (2025-03-01+)

DEFAULT_RATE = 120.0  # PHP/kWh — confirmed from MASTER


# ---------------------------------------------------------------------------
# Coercion helpers
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def coerce_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    if isinstance(value, str) and "VALUE" in value:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return default
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


# ---------------------------------------------------------------------------
# Row builder — creates one electricity_readings row per meter
# ---------------------------------------------------------------------------
def build_meter_row(
    reading_date: str,
    meter: str,
    start_kwh: float | None,
    end_kwh: float | None,
    diff: float | None,
    rate: float,
    source_row: int,
    section: str,  # "monthly" or "daily"
) -> dict[str, Any] | None:
    """
    Return a row dict or None if the row should be skipped (no meaningful data).
    For MAIN meter: skip if diff == 0 (not yet active month).
    For BUNKHOUSE/PUMP: skip if start_kwh and end_kwh are both 0/None.
    """
    # Normalize: treat 0 same as None for presence check
    has_start = start_kwh is not None and float(start_kwh) > 0
    has_end = end_kwh is not None and float(end_kwh) > 0

    # For BUNKHOUSE and PUMP in 2026 monthly section: all zeros mean not yet active
    if meter in ("BUNKHOUSE", "PUMP") and not has_start and not has_end:
        return None

    # For MAIN meter: if diff is 0 or None, skip (future month placeholder)
    if meter == "MAIN":
        if diff is None or float(diff) == 0:
            return None

    # March 2026 MAIN has start_kwh=0 (meter reset/replacement) — allow it,
    # it's a valid reading even if start is 0.
    effective_start = float(start_kwh) if start_kwh is not None else 0.0
    effective_end   = float(end_kwh)   if end_kwh   is not None else 0.0

    return {
        "reading_date": reading_date,
        "meter": meter,
        "start_kwh": effective_start,
        "end_kwh": effective_end,
        "rate_php_per_kwh": rate,
        "remarks": None,
        "_source_row": source_row,
        "_section": section,
    }


# ---------------------------------------------------------------------------
# Section A: Monthly summaries (rows 3-14)
# ---------------------------------------------------------------------------
def extract_monthly_section(ws, warnings: list[str]) -> list[dict[str, Any]]:
    rows = []
    for r in range(MONTHLY_DATA_START, MONTHLY_DATA_END + 1):
        raw_date = ws.cell(r, COL_DATE).value
        reading_date = coerce_date(raw_date)
        if reading_date is None:
            continue

        # MAIN meter
        total_start = coerce_float(ws.cell(r, COL_TOTAL_START).value)
        total_end   = coerce_float(ws.cell(r, COL_TOTAL_END).value)
        total_diff  = coerce_float(ws.cell(r, COL_TOTAL_DIFF).value)
        rate        = coerce_float(ws.cell(r, COL_RATE).value) or DEFAULT_RATE

        main_row = build_meter_row(
            reading_date, "MAIN", total_start, total_end, total_diff, rate, r, "monthly"
        )
        if main_row:
            if total_end is not None and total_start is not None and total_end < total_start:
                warnings.append(
                    f"ELEC monthly row {r} ({reading_date}): MAIN end_kwh ({total_end}) < start_kwh ({total_start})"
                )
            rows.append(main_row)

        # BUNKHOUSE meter
        bh_start = coerce_float(ws.cell(r, COL_BH_START).value)
        bh_end   = coerce_float(ws.cell(r, COL_BH_END).value)
        bh_diff  = coerce_float(ws.cell(r, COL_BH_DIFF).value)

        bh_row = build_meter_row(
            reading_date, "BUNKHOUSE", bh_start, bh_end, bh_diff, DEFAULT_RATE, r, "monthly"
        )
        if bh_row:
            rows.append(bh_row)

        # PUMP meter
        pump_start = coerce_float(ws.cell(r, COL_PUMP_START).value)
        pump_end   = coerce_float(ws.cell(r, COL_PUMP_END).value)
        pump_diff  = coerce_float(ws.cell(r, COL_PUMP_DIFF).value)

        pump_row = build_meter_row(
            reading_date, "PUMP", pump_start, pump_end, pump_diff, DEFAULT_RATE, r, "monthly"
        )
        if pump_row:
            rows.append(pump_row)

    return rows


# ---------------------------------------------------------------------------
# Section B: Daily detail (rows 19+)
# ---------------------------------------------------------------------------
def extract_daily_section(ws, warnings: list[str]) -> list[dict[str, Any]]:
    rows = []
    max_row = ws.max_row

    for r in range(DAILY_DATA_START, max_row + 1):
        raw_date = ws.cell(r, COL_DATE).value
        reading_date = coerce_date(raw_date)
        if reading_date is None:
            continue

        # MAIN (PLANT) meter in daily section — same column positions as monthly
        total_start = coerce_float(ws.cell(r, COL_TOTAL_START).value)
        total_end   = coerce_float(ws.cell(r, COL_TOTAL_END).value)
        total_diff  = coerce_float(ws.cell(r, COL_TOTAL_DIFF).value)

        main_row = build_meter_row(
            reading_date, "MAIN", total_start, total_end, total_diff, DEFAULT_RATE, r, "daily"
        )
        if main_row:
            if total_end is not None and total_start is not None and total_end < total_start:
                warnings.append(
                    f"ELEC daily row {r} ({reading_date}): MAIN end_kwh ({total_end}) < start_kwh ({total_start}) — FLAGGED"
                )
                main_row["_flagged"] = True
            rows.append(main_row)

        # BUNKHOUSE meter
        bh_start = coerce_float(ws.cell(r, COL_BH_START).value)
        bh_end   = coerce_float(ws.cell(r, COL_BH_END).value)
        bh_diff  = coerce_float(ws.cell(r, COL_BH_DIFF).value)

        bh_row = build_meter_row(
            reading_date, "BUNKHOUSE", bh_start, bh_end, bh_diff, DEFAULT_RATE, r, "daily"
        )
        if bh_row:
            rows.append(bh_row)

        # PUMP meter
        pump_start = coerce_float(ws.cell(r, COL_PUMP_START).value)
        pump_end   = coerce_float(ws.cell(r, COL_PUMP_END).value)
        pump_diff  = coerce_float(ws.cell(r, COL_PUMP_DIFF).value)

        pump_row = build_meter_row(
            reading_date, "PUMP", pump_start, pump_end, pump_diff, DEFAULT_RATE, r, "daily"
        )
        if pump_row:
            rows.append(pump_row)

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract electricity_readings from MASTER ELECTRICITY sheet."
    )
    parser.add_argument("--file", required=True, help="Path to MASTER ICTC INPUT FILE V1.xlsx")
    parser.add_argument("--output", default=None, help="Output JSON path (default: print to stdout)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    if "ELECTRICITY" not in wb.sheetnames:
        print(json.dumps({"error": "ELECTRICITY sheet not found", "available": list(wb.sheetnames)}), file=sys.stderr)
        return 2

    ws = wb["ELECTRICITY"]
    warnings: list[str] = []

    monthly_rows = extract_monthly_section(ws, warnings)
    daily_rows   = extract_daily_section(ws, warnings)

    all_rows = monthly_rows + daily_rows

    # Stats
    dates = [r["reading_date"] for r in all_rows]
    date_range = f"{min(dates)} -> {max(dates)}" if dates else "N/A"
    meters: dict[str, int] = {}
    for r in all_rows:
        m = r["meter"]
        meters[m] = meters.get(m, 0) + 1

    flagged = [r for r in all_rows if r.get("_flagged")]

    output = {
        "source_file": path.name,
        "source_sheet": "ELECTRICITY",
        "rows": all_rows,
        "summary": {
            "total_rows": len(all_rows),
            "monthly_rows": len(monthly_rows),
            "daily_rows": len(daily_rows),
            "flagged_rows": len(flagged),
            "date_range": date_range,
            "meters": meters,
            "total_warnings": len(warnings),
            "warnings": warnings,
        },
    }

    json_output = json.dumps(output, indent=2, default=str)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json_output)
        print(json.dumps({
            "ok": True,
            "total_rows": len(all_rows),
            "monthly_rows": len(monthly_rows),
            "daily_rows": len(daily_rows),
            "meters": meters,
            "warnings": len(warnings),
            "output": str(out_path),
        }))
    else:
        print(json_output)

    return 0


if __name__ == "__main__":
    sys.exit(main())
