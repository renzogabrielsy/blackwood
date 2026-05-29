#!/usr/bin/env python3
"""
MC's "Daily Production Report" extractor — one sheet per production DAY.

This is the most complex extractor in the sync-ictc pipeline: a single day-sheet
feeds FOUR target tables. It is a sibling of extract_proposed_daily.py (same coercion
helpers, same argparse/JSON shape) and reuses the field shapes documented in
extract_master_prod.py. Deterministic muscle only — NO DB access, NO network. JSON to stdout.

Source workbook (verified 2026-05-29 against sheets 05-27-26 and 05-28-26):
- One sheet per production day. Sheet title format "MM-DD-YY", almost always with heavy
  trailing whitespace (STRIP it). 2-digit year: 26 -> 2026.
- transaction_date = the SHEET TITLE date, NOT the in-sheet D4 header (D4 is the
  next-morning write date, e.g. sheet "05-27-26" has D4="MAY 28, 2026").

Scrape map (PRODUCTION_DESIGN.md Section 15.2):

  A. production_runs   — header row 7; data rows ~8-12
       D{r}=grade (customer prefix e.g. "CEBU 3X50") | E{r}=#sacks/bags | F{r}=#kilos/sack
       G{r}=TOTAL kg | H{r}=shift label ("MORNING SHIFT" / "NIGHT SHIFT")
       C13/G13 = TOTAL reconciliation row (CEBU-only day total) — captured in summary, not emitted.
       Routing: "CEBU <grade>" -> strip, customer="CEBU"; bare allowlist grade -> customer="CEBU".
       DROP (list, don't emit) KOREA POWDER / LOCAL POWDER / ZAMBOANGA ... and any grade
       not in {3X50, 6X50, 8X50, 2X6}.

  B. production_downtime — left block ~rows 24-27
       C24=category (e.g. "REPAIR") | C27=time-range(s) newline-sep | E27=minutes newline-sep
       F27=reason text. AGGREGATE per day: dt_mins=sum(minutes), dt_hrs=0,
       dt_reason=category + " | " + joined reasons, shift="M", shift_hrs=12 (default).
       Ignore the ambiguous C26 integer. Emit only when minutes>0 or a reason is present.

  C. electricity_readings — MAIN rows 53-60: D54=start_kwh, E54=end_kwh, E60=meter_multiplier(120).
       BUNKHOUSE row 65 + PUMP row 67: D=start, E=end. Emit only when non-blank/non-zero.
       DB has generated consumption_kwh = (end-start)*meter_multiplier — do NOT emit consumption.

  D. truck_readings — header row 46; data rows 47, 49, 51
       C{r}=plate | D{r}=start_km | E{r}=end_km | F{r}=total distance | H{r}=liters issued
       J{r}/K{r}=qualitative fuel gauge -> folded into remarks.
       Emit only when end_km>start_km OR F>0 OR fuel present. Skip fully idle rows.

Usage:
    python3 extract_daily_production.py --file path/to/Daily Production Report.xlsx
    python3 extract_daily_production.py --file "..." --sheet "05-27-26"
    python3 extract_daily_production.py --file "..." --all-sheets
    python3 extract_daily_production.py --file "..." --year 2026   # optional century override
    python3 extract_daily_production.py --file "..." --all-sheets --since 2026-05-23
        # ^ keep ONLY day-sheets dated STRICTLY AFTER the watermark (exclusive). The
        #   watermark is the latest already-ingested date, which we do NOT re-ingest.
        #   Omit --since for full-history backfill (identical to today's behavior).

Output: JSON to stdout — see the module-level docstring of main() for the full shape.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)


# ---------------------------------------------------------------------------
# Domain constants
# ---------------------------------------------------------------------------
# Finished-grade allowlist for production_runs (PRODUCTION_DESIGN.md §3, §15.2).
# Anything not in this set (KOREA POWDER, LOCAL POWDER, ZAMBOANGA ...) is dropped.
VALID_GRADES = {"3X50", "6X50", "8X50", "2X6"}

# Shift label -> canonical DB code (PRODUCTION_DESIGN.md §15.6).
# NOTE: "NIGHT SHIFT" maps to E (the 2nd shift's canonical code in this DB), NOT N.
SHIFT_LABEL_TO_CODE = {
    "MORNING SHIFT": "M",
    "MORNING": "M",
    "NIGHT SHIFT": "E",
    "NIGHT": "E",
    "EVENING SHIFT": "E",
    "EVENING": "E",
    "AFTERNOON SHIFT": "E",
}

# production_batch = full English month name UPPERCASE derived from transaction_date.
MONTH_NAME_UPPER = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL", 5: "MAY", 6: "JUNE",
    7: "JULY", 8: "AUGUST", 9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER",
}

# Sheet title "MM-DD-YY" (tolerate any surrounding whitespace; the strip in
# parse_sheet_date handles it, but the regex stays permissive too).
SHEET_NAME_RE = re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{2,4})$")

# ---- Cell coordinates (verified on 05-27-26 / 05-28-26) -------------------
# Section A — production output
RUNS_HEADER_ROW = 7
RUNS_FIRST_DATA_ROW = 8
RUNS_LAST_DATA_ROW = 12       # rows 8..12 inclusive scanned for grades
COL_RUN_GRADE = 4             # D
COL_RUN_SACKS = 5             # E
COL_RUN_KILOS_PER = 6        # F
COL_RUN_TTL_KG = 7           # G
COL_RUN_SHIFT = 8            # H
TOTAL_ROW = 13               # C13="TOTAL", G13=day total

# Section B — downtime (left block)
COL_DT_CATEGORY = 3          # C24
DT_CATEGORY_ROW = 24
DT_RANGES_ROW = 27           # C27 time ranges
DT_MINUTES_ROW = 27          # E27 minutes
DT_REASON_ROW = 27           # F27 reasons
COL_DT_RANGES = 3            # C
COL_DT_MINUTES = 5           # E
COL_DT_REASON = 6            # F

# Section C — electricity
ELEC_MAIN_READING_ROW = 54   # D54=start, E54=end
ELEC_MAIN_MULT_ROW = 60      # E60=multiplier (120); E59 is the header
COL_ELEC_START = 4           # D
COL_ELEC_END = 5             # E
COL_ELEC_MULT = 5            # E (on the multiplier row)
ELEC_BUNKHOUSE_ROW = 65      # A65="BUNKHOUSE"
ELEC_PUMP_ROW = 67           # A67="PUMP"
DEFAULT_METER_MULTIPLIER = 120.0

# Section D — trucks
TRUCK_HEADER_ROW = 46
TRUCK_DATA_ROWS = (47, 49, 51)
COL_TRUCK_PLATE = 3          # C
COL_TRUCK_START_KM = 4       # D
COL_TRUCK_END_KM = 5         # E
COL_TRUCK_TTL_KM = 6         # F
COL_TRUCK_LITERS = 8         # H
COL_TRUCK_GAUGE_START = 10   # J
COL_TRUCK_GAUGE_END = 11     # K


# ---------------------------------------------------------------------------
# Coercion helpers (mirrors extract_proposed_daily.py)
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> int | None:
    f = coerce_float(value)
    return int(f) if f is not None else None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Sheet-name / date helpers
# ---------------------------------------------------------------------------
def parse_sheet_date(sheet_name: str, year_override: int | None) -> date | None:
    """
    Parse a "MM-DD-YY" sheet title into a date. Tolerates trailing/leading whitespace.
    2-digit year maps via the 2000-century rule (26 -> 2026). year_override, if given,
    forces the year regardless of the sheet's YY.
    """
    m = SHEET_NAME_RE.match(sheet_name.strip())
    if not m:
        return None
    month = int(m.group(1))
    day = int(m.group(2))
    yy = m.group(3)
    if year_override is not None:
        year = year_override
    elif len(yy) == 2:
        year = 2000 + int(yy)
    else:
        year = int(yy)
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_since(value: str | None) -> date | None:
    """
    Parse the optional --since watermark ('YYYY-MM-DD') into a date. Returns None when
    --since is omitted (no filtering). Raises ValueError on a malformed value so main()
    can surface a clean error instead of silently extracting everything.
    """
    if value is None:
        return None
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def production_batch_for(d: date) -> str:
    return MONTH_NAME_UPPER[d.month]


def normalize_shift(label: Any) -> tuple[str | None, str | None]:
    """Return (code, warning). code is M/E/N or None if unrecognized."""
    s = coerce_str(label)
    if s is None:
        return None, None
    code = SHIFT_LABEL_TO_CODE.get(s.upper())
    if code is None:
        return None, f"Unrecognized shift label '{s}'"
    return code, None


def split_multiline(value: Any) -> list[str]:
    """Split a newline-separated cell into non-empty trimmed lines (skips blank spacer lines)."""
    s = coerce_str(value)
    if s is None:
        return []
    return [line.strip() for line in s.splitlines() if line.strip()]


# ---------------------------------------------------------------------------
# Section A — production runs
# ---------------------------------------------------------------------------
def route_grade(raw_grade: str) -> tuple[str | None, str | None, bool]:
    """
    Map MC's grade cell text to (customer, grade, keep).

    The cell is "<CUSTOMER> <GRADE>" or a bare "<GRADE>". A finished-grade run is kept
    iff the trailing token is a valid mesh size in VALID_GRADES; the leading token(s)
    become the customer. Verified grade-cell values across the live workbook:
      - "CEBU 3X50"            -> ("CEBU",    "3X50", True)   # implicit-customer day
      - "KURARAY 3X50"         -> ("KURARAY", "3X50", True)   # real customer crossover (DESIGN §12/§3)
      - bare "6X50"            -> ("CEBU",    "6X50", True)   # bare grade defaults to CEBU
      - "KOREA POWDER (BAGGED)"-> dropped (trailing token not a grade)  # waste-sale, out of v1 scope
      - "LOCAL POWDER" / "ZAMBOANGA ..." -> dropped (no valid grade token)

    Routing is intentionally driven by the GRADE allowlist, not a CEBU-only customer
    allowlist: dropping a legitimate KURARAY finished-grade run would silently lose real
    production tonnage. The schema's production_runs.customer column exists precisely for
    this (default 'CEBU'; KURARAY observed in the MASTER backfill).
    """
    text = raw_grade.strip().upper()
    tokens = text.split()
    if not tokens:
        return None, None, False

    # Bare single-token grade -> implicit CEBU customer.
    if len(tokens) == 1:
        grade = tokens[0]
        return ("CEBU", grade, True) if grade in VALID_GRADES else (None, None, False)

    # Multi-token: trailing token is the candidate grade, the rest is the customer.
    grade = tokens[-1]
    customer = " ".join(tokens[:-1])
    if grade in VALID_GRADES:
        return customer, grade, True
    return None, None, False


def extract_runs(
    ws,
    txn_date: date,
    production_batch: str,
    sheet_warnings: list[str],
    dropped_grades: list[str],
) -> tuple[list[dict[str, Any]], float | None]:
    """Extract production_runs rows. Returns (runs, day_total_g13)."""
    runs: list[dict[str, Any]] = []

    for r in range(RUNS_FIRST_DATA_ROW, RUNS_LAST_DATA_ROW + 1):
        raw_grade = coerce_str(ws.cell(r, COL_RUN_GRADE).value)
        if raw_grade is None:
            continue  # blank separator row

        customer, grade, keep = route_grade(raw_grade)
        if not keep:
            dropped_grades.append(raw_grade)
            continue

        ttl_kg = coerce_float(ws.cell(r, COL_RUN_TTL_KG).value)
        sacks = coerce_int(ws.cell(r, COL_RUN_SACKS).value)
        shift_code, shift_warn = normalize_shift(ws.cell(r, COL_RUN_SHIFT).value)

        row_warnings: list[str] = []
        if shift_warn:
            row_warnings.append(shift_warn)
        if shift_code is None:
            row_warnings.append(f"missing/invalid shift for grade {grade}")
        if ttl_kg is None:
            row_warnings.append(f"missing TOTAL kg (G{r}) for {grade}")
        elif ttl_kg < 0:
            row_warnings.append(f"negative ttl_kg={ttl_kg}")

        for w in row_warnings:
            sheet_warnings.append(f"[{ws.title.strip()}] runs R{r}: {w}")

        confidence = max(0.0, 1.0 - 0.10 * len(row_warnings))

        runs.append({
            "transaction_date": txn_date.isoformat(),
            "production_batch": production_batch,
            "shift": shift_code,
            "customer": customer,
            "grade": grade,
            "ttl_kg": ttl_kg,
            "sacks_bags": sacks,
            "remarks": None,
            "_source_sheet": ws.title.strip(),
            "_source_row": r,
            "warnings": row_warnings,
            "confidence": round(confidence, 3),
        })

    # G13 reconciliation total (CEBU-only day total). Only trust it when C13 says TOTAL.
    day_total = None
    c13 = coerce_str(ws.cell(TOTAL_ROW, COL_RUN_GRADE - 1).value)  # C13
    g13 = coerce_float(ws.cell(TOTAL_ROW, COL_RUN_TTL_KG).value)   # G13
    if c13 and c13.strip().upper() == "TOTAL":
        day_total = g13
    elif g13 is not None:
        # Fall back to G13 even if the label is missing — but flag it.
        day_total = g13

    return runs, day_total


# ---------------------------------------------------------------------------
# Section B — downtime (aggregated to one row on the M shift)
# ---------------------------------------------------------------------------
def extract_downtime(
    ws,
    txn_date: date,
    production_batch: str,
    sheet_warnings: list[str],
) -> dict[str, Any] | None:
    """Aggregate the day's downtime events into a single M-shift row, or None if no downtime."""
    category = coerce_str(ws.cell(DT_CATEGORY_ROW, COL_DT_CATEGORY).value)
    ranges = split_multiline(ws.cell(DT_RANGES_ROW, COL_DT_RANGES).value)
    minute_lines = split_multiline(ws.cell(DT_MINUTES_ROW, COL_DT_MINUTES).value)
    reasons = split_multiline(ws.cell(DT_REASON_ROW, COL_DT_REASON).value)

    row_warnings: list[str] = []

    # Parse minutes — each line looks like "9 MINUTES" / "19 MINUTES".
    total_mins = 0.0
    parsed_any = False
    for line in minute_lines:
        f = coerce_float(re.sub(r"[^0-9.]", "", line))
        if f is not None:
            total_mins += f
            parsed_any = True
        else:
            row_warnings.append(f"could not parse downtime minutes from '{line}'")

    has_reason = bool(reasons) or bool(category)

    # Only emit when there's real downtime: minutes > 0 OR a reason/category present.
    if total_mins <= 0 and not has_reason:
        return None

    # Build dt_reason = category + " | " + joined reason lines.
    reason_parts: list[str] = []
    if category:
        reason_parts.append(category)
    if reasons:
        reason_parts.append("; ".join(reasons))
    dt_reason = " | ".join(reason_parts) if reason_parts else None

    if total_mins <= 0 and has_reason:
        row_warnings.append("downtime reason present but no parseable minutes")

    # Fold the time-range strings into remarks for traceability.
    remarks = ("Time ranges: " + "; ".join(ranges)) if ranges else None

    for w in row_warnings:
        sheet_warnings.append(f"[{ws.title.strip()}] downtime: {w}")

    return {
        "transaction_date": txn_date.isoformat(),
        "production_batch": production_batch,
        "shift": "M",            # locked decision: attach aggregated downtime to Morning
        "shift_hrs": 12,         # default — email has no clean shift-length value (ignores C26)
        "dt_hrs": 0,
        "dt_mins": total_mins,
        "dt_reason": dt_reason,
        "remarks": remarks,
        "_source_sheet": ws.title.strip(),
        "warnings": row_warnings,
    }


# ---------------------------------------------------------------------------
# Section C — electricity
# ---------------------------------------------------------------------------
def _emit_electricity(
    meter: str,
    start_kwh: float | None,
    end_kwh: float | None,
    multiplier: float | None,
    txn_date: date,
    ws_title: str,
    sheet_warnings: list[str],
) -> dict[str, Any] | None:
    """Build an electricity row, or None if the meter is blank/idle (no readings)."""
    # Idle/blank: both readings missing, or both zero.
    if start_kwh is None and end_kwh is None:
        return None
    if (start_kwh or 0) == 0 and (end_kwh or 0) == 0:
        return None

    row_warnings: list[str] = []
    if start_kwh is None:
        row_warnings.append("missing start_kwh")
    if end_kwh is None:
        row_warnings.append("missing end_kwh")
    if start_kwh is not None and end_kwh is not None and end_kwh < start_kwh:
        row_warnings.append(f"end_kwh ({end_kwh}) < start_kwh ({start_kwh})")

    mult = multiplier if multiplier is not None else DEFAULT_METER_MULTIPLIER

    for w in row_warnings:
        sheet_warnings.append(f"[{ws_title.strip()}] electricity {meter}: {w}")

    return {
        "reading_date": txn_date.isoformat(),
        "meter": meter,
        "start_kwh": start_kwh,
        "end_kwh": end_kwh,
        "meter_multiplier": mult,
        "remarks": None,
        "_source_sheet": ws_title.strip(),
        "warnings": row_warnings,
    }


def extract_electricity(
    ws,
    txn_date: date,
    sheet_warnings: list[str],
) -> list[dict[str, Any]]:
    readings: list[dict[str, Any]] = []

    # MAIN
    main_start = coerce_float(ws.cell(ELEC_MAIN_READING_ROW, COL_ELEC_START).value)
    main_end = coerce_float(ws.cell(ELEC_MAIN_READING_ROW, COL_ELEC_END).value)
    main_mult = coerce_float(ws.cell(ELEC_MAIN_MULT_ROW, COL_ELEC_MULT).value)
    main = _emit_electricity(
        "MAIN", main_start, main_end, main_mult, txn_date, ws.title, sheet_warnings
    )
    if main is not None:
        readings.append(main)

    # BUNKHOUSE + PUMP (idle in 2026 — usually skipped). Multiplier defaults to 120.
    for meter, row in (("BUNKHOUSE", ELEC_BUNKHOUSE_ROW), ("PUMP", ELEC_PUMP_ROW)):
        start = coerce_float(ws.cell(row, COL_ELEC_START).value)
        end = coerce_float(ws.cell(row, COL_ELEC_END).value)
        rec = _emit_electricity(
            meter, start, end, None, txn_date, ws.title, sheet_warnings
        )
        if rec is not None:
            readings.append(rec)

    return readings


# ---------------------------------------------------------------------------
# Section D — trucks
# ---------------------------------------------------------------------------
def extract_trucks(
    ws,
    txn_date: date,
    sheet_warnings: list[str],
) -> list[dict[str, Any]]:
    trucks: list[dict[str, Any]] = []

    for r in TRUCK_DATA_ROWS:
        plate = coerce_str(ws.cell(r, COL_TRUCK_PLATE).value)
        start_km = coerce_float(ws.cell(r, COL_TRUCK_START_KM).value)
        end_km = coerce_float(ws.cell(r, COL_TRUCK_END_KM).value)
        ttl_km = coerce_float(ws.cell(r, COL_TRUCK_TTL_KM).value)
        fuel = coerce_float(ws.cell(r, COL_TRUCK_LITERS).value)
        gauge_start = coerce_str(ws.cell(r, COL_TRUCK_GAUGE_START).value)
        gauge_end = coerce_str(ws.cell(r, COL_TRUCK_GAUGE_END).value)

        # Determine movement: end>start, or a positive total distance, or fuel issued.
        moved = (
            (start_km is not None and end_km is not None and end_km > start_km)
            or (ttl_km is not None and ttl_km > 0)
        )
        has_fuel = fuel is not None and fuel > 0

        if not moved and not has_fuel:
            continue  # fully idle row — skip (plate may also be blank, e.g. forklift slot)

        row_warnings: list[str] = []
        if plate is None:
            row_warnings.append("movement/fuel present but plate is blank")

        # Fold the qualitative fuel gauge into remarks.
        gauge_bits = []
        if gauge_start:
            gauge_bits.append(f"start fuel: {gauge_start}")
        if gauge_end:
            gauge_bits.append(f"arriving fuel: {gauge_end}")
        remarks = "; ".join(gauge_bits) if gauge_bits else None

        for w in row_warnings:
            sheet_warnings.append(f"[{ws.title.strip()}] trucks R{r}: {w}")

        trucks.append({
            "reading_date": txn_date.isoformat(),
            "plate_no": plate,
            "start_km": start_km,
            "end_km": end_km,
            "fuel_liters": fuel if has_fuel else None,
            "remarks": remarks,
            "_source_sheet": ws.title.strip(),
            "_source_row": r,
            "warnings": row_warnings,
        })

    return trucks


# ---------------------------------------------------------------------------
# Per-sheet orchestration
# ---------------------------------------------------------------------------
def extract_sheet(
    ws,
    year_override: int | None,
) -> dict[str, Any]:
    """Extract all four record types from one day-sheet."""
    sheet_warnings: list[str] = []
    txn_date = parse_sheet_date(ws.title, year_override)
    if txn_date is None:
        sheet_warnings.append(
            f"Sheet '{ws.title.strip()}': cannot parse MM-DD-YY date from title"
        )
        return {
            "transaction_date": None,
            "runs": [],
            "downtime": [],
            "electricity": [],
            "trucks": [],
            "day_total_g13": None,
            "dropped_grades": [],
            "warnings": sheet_warnings,
        }

    production_batch = production_batch_for(txn_date)
    dropped_grades: list[str] = []

    runs, day_total = extract_runs(
        ws, txn_date, production_batch, sheet_warnings, dropped_grades
    )
    downtime = extract_downtime(ws, txn_date, production_batch, sheet_warnings)
    electricity = extract_electricity(ws, txn_date, sheet_warnings)
    trucks = extract_trucks(ws, txn_date, sheet_warnings)

    return {
        "transaction_date": txn_date.isoformat(),
        "runs": runs,
        "downtime": [downtime] if downtime is not None else [],
        "electricity": electricity,
        "trucks": trucks,
        "day_total_g13": day_total,
        "dropped_grades": dropped_grades,
        "warnings": sheet_warnings,
    }


# ---------------------------------------------------------------------------
# Sheet selection
# ---------------------------------------------------------------------------
def resolve_sheets(wb, args, since: date | None) -> tuple[list[str], str | None]:
    """
    Return (sheet_names_to_process, error). Matching tolerates trailing whitespace in
    workbook sheet titles.

    A day-sheet's date applies to ALL of its record types (runs/downtime/electricity/
    trucks share the sheet's MM-DD-YY title date), so the --since watermark is applied
    at the sheet level: whole sheets dated <= since are dropped here. This is the
    cleanest place to filter — every emitted record from a kept sheet is guaranteed
    newer than the watermark, and the summary stays internally consistent because it
    only ever sees kept sheets. Sheets whose title date cannot be parsed are KEPT (so
    extract_sheet can surface the parse warning) — they emit no dated records anyway.
    """
    if args.all_sheets:
        selected = list(wb.sheetnames)
    elif args.sheet:
        target = args.sheet.strip()
        matches = [n for n in wb.sheetnames if n.strip() == target]
        if not matches:
            return [], (
                f"Sheet '{target}' not found. Available (stripped): "
                f"{[n.strip() for n in wb.sheetnames]}"
            )
        selected = matches
    else:
        # Default: the LATEST day-sheet (max parseable MM-DD-YY date). Falls back to the
        # last sheet in the workbook if none parse.
        dated = []
        for n in wb.sheetnames:
            d = parse_sheet_date(n, args.year)
            if d is not None:
                dated.append((d, n))
        if dated:
            dated.sort(key=lambda t: t[0])
            selected = [dated[-1][1]]
        else:
            selected = [wb.sheetnames[-1]]

    # Apply the --since watermark (exclusive): keep sheets dated STRICTLY AFTER since.
    # Unparseable-title sheets are kept (they emit no dated records and carry a warning).
    if since is not None:
        kept = []
        for n in selected:
            d = parse_sheet_date(n, args.year)
            if d is None or d > since:
                kept.append(n)
        selected = kept

    return selected, None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract production runs/downtime/electricity/trucks from MC's Daily Production Report."
    )
    parser.add_argument("--file", required=True, help="Path to the Daily Production Report XLSX")
    parser.add_argument("--year", type=int, default=None,
                        help="Optional century override for sheet-name YY (e.g., 2026). "
                             "Default: derive from the 2-digit year in each sheet title.")
    parser.add_argument("--sheet", help="Specific sheet name 'MM-DD-YY' (trailing whitespace ok). "
                                        "Default: latest sheet.")
    parser.add_argument("--all-sheets", action="store_true",
                        help="Process every sheet (catch-up/backfill).")
    parser.add_argument("--since", default=None,
                        help="Watermark date 'YYYY-MM-DD'. Keep ONLY day-sheets dated "
                             "STRICTLY AFTER this date (exclusive — the watermark is the "
                             "latest already-ingested date, which is NOT re-emitted). "
                             "Omit for full-history backfill (no filtering).")
    args = parser.parse_args()

    try:
        since = parse_since(args.since)
    except ValueError:
        print(json.dumps({
            "error": f"Invalid --since '{args.since}'. Expected format YYYY-MM-DD."
        }), file=sys.stderr)
        return 2

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    sheet_names, err = resolve_sheets(wb, args, since)
    if err:
        print(json.dumps({"error": err}), file=sys.stderr)
        return 2

    all_runs: list[dict] = []
    all_downtime: list[dict] = []
    all_electricity: list[dict] = []
    all_trucks: list[dict] = []
    all_warnings: list[str] = []
    sheets_processed: list[str] = []
    dropped_grades: list[str] = []
    day_totals: dict[str, float | None] = {}

    # Infer the "primary" year for the top-level output field.
    inferred_years: set[int] = set()

    for name in sheet_names:
        ws = wb[name]
        result = extract_sheet(ws, args.year)
        all_runs.extend(result["runs"])
        all_downtime.extend(result["downtime"])
        all_electricity.extend(result["electricity"])
        all_trucks.extend(result["trucks"])
        all_warnings.extend(result["warnings"])
        dropped_grades.extend(result["dropped_grades"])
        sheets_processed.append(name.strip())
        if result["transaction_date"]:
            day_totals[result["transaction_date"]] = result["day_total_g13"]
            inferred_years.add(int(result["transaction_date"][:4]))

    if args.year is not None:
        year_out: int | None = args.year
    elif len(inferred_years) == 1:
        year_out = next(iter(inferred_years))
    else:
        year_out = None  # mixed or none

    # Overall confidence: mean of per-run confidences (the only scored record type).
    confs = [r["confidence"] for r in all_runs if r.get("confidence") is not None]
    overall_conf = round(sum(confs) / len(confs), 3) if confs else 1.0

    output = {
        "filename": path.name,
        "year": year_out,
        "sheets_processed": sheets_processed,
        "runs": all_runs,
        "downtime": all_downtime,
        "electricity": all_electricity,
        "trucks": all_trucks,
        "summary": {
            "sheets": len(sheets_processed),
            "runs_count": len(all_runs),
            "downtime_count": len(all_downtime),
            "electricity_count": len(all_electricity),
            "trucks_count": len(all_trucks),
            "dropped_grades": dropped_grades,
            "day_totals": day_totals,
            "extraction_warnings": all_warnings,
            "overall_confidence": overall_conf,
        },
    }

    print(json.dumps(output, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
