#!/usr/bin/env python3
"""
RC DELIVERIES XLSX extractor — v2, matches the operator's actual sheet format.

Real-world format discovered on 2026-05-26 working with sample files from
Ivy + Pretchel:

- Workbook has one sheet per MONTH (e.g. "JANUARY 2026", ..., "MAY 2026")
- Row 1 = title only ("MAY 2026 DELIVERIES")
- Row 2 = main headers (DATE, DATE OF, Sample, Block, block, TRUCK, ...)
- Row 3 = sub-headers (ANALYZED, DELIVERY, Supplier name, locator, ...)
- Row 4-5 = spec thresholds (8.9 max., 0.575 min., etc.)
- Row 6+ = data rows
- Sparse date cells — only first row of each day has a date; continuations
  are blank and must be forward-filled
- "Block" column (col 4) holds the OPERATOR's batch label like "B09",
  not Blackwood's full batch_code. We translate heuristically using
  the Remarks column when possible ("PILED IN APRIL # 9" -> "APRIL-26-BLK9")
- No price columns in the operator's file — cost_basis is always null here
- Trailing rows: blank rows + an "Average" summary row at the bottom

Usage:
    python3 extract_rc_deliveries.py --file path/to/RC_DELIVERIES.xlsx
    python3 extract_rc_deliveries.py --file path/to/RC_DELIVERIES.xlsx --all-sheets
    python3 extract_rc_deliveries.py --file path/to/RC_DELIVERIES.xlsx --sheet "MAY 2026"

Output:
    {
      "filename": "...",
      "sheets_processed": ["MAY 2026"],
      "rows": [ {row_dict}, ... ],
      "summary": {
        "total_rows": N,
        "extraction_warnings": [...],
        "overall_confidence": 0.0-1.0,
        "unmapped_batches": ["B09", ...]   # operator labels that couldn't be heuristically translated
      }
    }
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)

# Shared weight-deduction grammar + wet-recovery core (one source of truth shared
# with extract_gsheet.py — see DEDUCTIONS_DESIGN.md / LEARNING_LEDGER L-021).
# When this script is run directly, sys.path[0] is its own dir, so `lib` (a sibling
# package with __init__.py) imports the same way lib.db is imported elsewhere.
from lib.deductions import (  # noqa: E402
    detect_deduction,
    build_recovery_row,
    is_recovery_row_dict,
    _is_inheritable_mother,
)


# ---------------------------------------------------------------------------
# Operator-specific format constants
# ---------------------------------------------------------------------------
# Column index (1-based) -> canonical field name
# Based on the format observed in real RC DELIVERIES 2026.xlsx (2026-05).
OPERATOR_COLUMNS: dict[int, str] = {
    1: "date_analyzed",       # informational, sometimes differs from transaction_date
    2: "transaction_date",
    3: "supplier",
    4: "operator_batch_label",  # raw "Block" value, e.g. "B09", "FEEDING AREA 1"
    5: "block_loc",
    6: "truck_plate",
    7: "weight_kg",
    8: "sacks",
    9: "lab_mc",
    10: "lab_grit",
    11: "lab_bd_astm",
    12: "lab_bd_jis",
    13: "lab_vm",
    14: "lab_ash",
    15: "lab_fc",
    16: "remarks",
}

# Validation thresholds (Blackwood schema rules — see CLAUDE.md)
WEIGHT_KG_MIN = 0
WEIGHT_KG_MAX = 100_000

BLOCK_LOC_REGEX = re.compile(r"^(PCA|PCB|[A-DF])-\d{1,2}[A-D]$")

LAB_PLAUSIBILITY: dict[str, tuple[str, Any]] = {
    "mc":  ("Moisture content unusually high",   lambda v: v < 20),
    "ash": ("Ash content unusually high",        lambda v: v < 10),
    "fc":  ("Fixed carbon unusually low",        lambda v: v > 60),
    "vm":  ("Volatile matter unusually high",    lambda v: v < 25),
    "grit": ("Grit value unusually high",        lambda v: v < 5),
    "bd_astm": ("BD ASTM out of expected range", lambda v: 0.2 < v < 1.0),
    "bd_jis":  ("BD JIS out of expected range",  lambda v: 0.2 < v < 1.0),
}

# Month-name -> Blackwood batch_code month prefix.
# Empirically verified from the DB: APRIL-26-BLK*, MAY-26-BLK*, etc.
# All months use the FULL name (MAY is naturally 3 letters).
MONTH_ABBR = {
    "JANUARY": "JANUARY", "FEBRUARY": "FEBRUARY", "MARCH": "MARCH",
    "APRIL": "APRIL", "MAY": "MAY", "JUNE": "JUNE", "JULY": "JULY",
    "AUGUST": "AUGUST", "SEPTEMBER": "SEPTEMBER", "OCTOBER": "OCTOBER",
    "NOVEMBER": "NOVEMBER", "DECEMBER": "DECEMBER",
}

# Heuristic: remarks containing "PILED IN <MONTH> # <N>" -> batch_code
PILED_REMARK_RE = re.compile(
    r"PILED\s+IN\s+(?P<month>JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s*#\s*(?P<num>\d+)",
    re.IGNORECASE,
)

# Operator label patterns
OPERATOR_B_LABEL_RE = re.compile(r"^B0?(\d{1,3})$", re.IGNORECASE)
# A FEEDING-area label in the operator's Block column (column D).
#
# WIDENED 2026-08-13 (L-042). It used to be r"^FEEDING\s+AREA\s*(\d*)$", which matched the
# spelling the SHEET uses and NOT the one MC actually types. She writes "FEEDING # 1", so the
# label fell through to the raw-value branch: truthy (so it passed the malformed guard), not
# pattern-valid (so it never auto-created), and therefore held on EVERY run forever.
#
# ACCEPTED (all produce EXACTLY what "FEEDING AREA <N>" produces today):
#   FEEDING AREA 2 / FEEDING # 2 / FEEDING #2 / FEEDING NO. 2 / FEEDING NO 2 / FEEDING 2 /
#   FEEDING AREA #2 / FEEDING AREA 2. / bare FEEDING (numberless -> raw label + warning).
# STILL REJECTED: FEEDING AREA A, FEEDING AREA 1 AND 2, RE-FEEDING 1, FEEDINGS 2.
#
# MUST stay byte-equivalent to workers/sync/src/reports/deliveries/extract.ts.
FEEDING_AREA_RE = re.compile(r"^FEEDING(?:\s*(?:AREA|NO))?\s*[#.:-]?\s*(\d*)\s*\.?$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Coercion helpers
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> str | None:
    """Return YYYY-MM-DD or None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):  # bool is a subtype of int — reject
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("₱", "").replace("$", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> int | None:
    f = coerce_float(value)
    return int(f) if f is not None else None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def find_header_row(sheet) -> int | None:
    """
    Locate the row that contains the operator's main header line.
    Heuristic: the row has "DATE OF" or "DELIVERY" in cols 1-3 AND "Supplier"
    or "Sample" somewhere in cols 2-5.
    Returns the row number (1-based) of the header row, or None if not found.
    """
    for r in range(1, min(sheet.max_row + 1, 16)):
        # Concatenate cells in cols 1-6 into a single search string
        first6 = " ".join(
            str(sheet.cell(r, c).value or "").upper()
            for c in range(1, 7)
        )
        if (
            ("DATE OF" in first6 or "DELIVERY" in first6)
            and ("SUPPLIER" in first6 or "SAMPLE" in first6)
        ):
            return r
    return None


def first_data_row_below(sheet, header_row: int) -> int:
    """
    Find the first row of actual data below the header.
    Operator's format: header (R2) + sub-header (R3) + 1-2 spec rows (R4-R5).
    Data starts at the first row below header_row whose col 2 (transaction_date)
    parses as a date.
    """
    # Scan up to 8 rows past header for a date cell
    for r in range(header_row + 1, min(sheet.max_row + 1, header_row + 9)):
        if coerce_date(sheet.cell(r, 2).value) is not None:
            return r
    # Fallback: header_row + 4 (skip header + sub-header + 2 spec rows)
    return header_row + 4


# ---------------------------------------------------------------------------
# Batch code translation
# ---------------------------------------------------------------------------
def translate_batch_code(
    operator_label: str | None,
    remarks: str | None,
    delivery_date: str | None,
) -> tuple[str | None, list[str]]:
    """
    Heuristically translate the operator's batch label into a Blackwood
    batch_code. Returns (batch_code, warnings).

    Rules in priority order:
      1. Remarks contains "PILED IN <MONTH> # <N>" -> "<MMM>-<YY>-BLK<N>"
      2. operator_label matches "B<N>" + delivery_date known
         -> "<MMM_OF_DELIVERY_MONTH>-<YY>-BLK<N>"
      3. operator_label matches "FEEDING AREA X" -> "FEED"
      4. Otherwise -> raw operator_label, warning added
    """
    warnings: list[str] = []
    if not operator_label:
        return None, ["No operator batch label in row"]

    label = operator_label.strip()

    # Rule 3: a FEEDING label (FEEDING AREA N / FEEDING # N / ...) ->
    # "<MMM>-<YY>-FEED<N>" (month from delivery_date). ONE output shape for every spelling.
    m_feed = FEEDING_AREA_RE.match(label)
    if m_feed:
        feed_num = m_feed.group(1)
        if feed_num and delivery_date:
            try:
                dt = datetime.strptime(delivery_date, "%Y-%m-%d")
                mmm = list(MONTH_ABBR.values())[dt.month - 1]
                yy = delivery_date[2:4]
                return f"{mmm}-{yy}-FEED{int(feed_num)}", []
            except Exception:
                pass
        # No number or no date — return raw with warning so user can review
        return label, [
            f"FEEDING label '{label}' could not be auto-numbered "
            f"(missing area number or delivery date). Needs manual mapping."
        ]

    # Rule 1: PILED IN <MONTH> # <N> in remarks
    if remarks:
        m = PILED_REMARK_RE.search(remarks)
        if m:
            month_name = m.group("month").upper()
            num = int(m.group("num"))
            mmm = MONTH_ABBR.get(month_name)
            # YY: use delivery year if available, else current year
            yy = "26"
            if delivery_date:
                try:
                    yy = delivery_date[2:4]
                except Exception:
                    pass
            if mmm:
                return f"{mmm}-{yy}-BLK{num}", []

    # Rule 2: B<N> label, infer month from delivery_date
    m = OPERATOR_B_LABEL_RE.match(label)
    if m and delivery_date:
        num = int(m.group(1))
        try:
            dt = datetime.strptime(delivery_date, "%Y-%m-%d")
            mmm_index = dt.month  # 1-12
            mmm = list(MONTH_ABBR.values())[mmm_index - 1]
            yy = delivery_date[2:4]
            return f"{mmm}-{yy}-BLK{num}", [
                f"Batch code translated heuristically: '{label}' -> '{mmm}-{yy}-BLK{num}' "
                f"(no remarks hint, used delivery month {mmm})"
            ]
        except Exception:
            pass

    # Fallthrough: emit raw label with warning
    return label, [
        f"Could not map operator batch label '{label}' to a Blackwood batch_code; "
        f"emitting raw value. Row may need manual mapping."
    ]


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------
def is_average_or_summary_row(sheet, row_num: int) -> bool:
    """The operator file ends with rows like 'Average', '100', etc. — skip them."""
    col1 = sheet.cell(row_num, 1).value
    if isinstance(col1, str) and col1.strip().lower() in {"average", "total", "sum"}:
        return True
    # Row with no supplier AND no weight AND no date is a noise row
    if (
        coerce_str(sheet.cell(row_num, 3).value) is None
        and coerce_float(sheet.cell(row_num, 7).value) is None
        and coerce_date(sheet.cell(row_num, 2).value) is None
    ):
        return True
    return False


def extract_row(
    sheet,
    row_num: int,
    last_seen_date: str | None,
) -> tuple[dict[str, Any] | None, str | None, list[str]]:
    """
    Extract one row. Returns (row_dict | None, new_last_seen_date, warnings).

    Returns None for row_dict if the row should be skipped (Average, fully blank, no usable data).
    `new_last_seen_date` is the date to forward-fill into subsequent blank-date rows.
    """
    warnings: list[str] = []

    if is_average_or_summary_row(sheet, row_num):
        return None, last_seen_date, []

    # Forward-fill date
    raw_date = sheet.cell(row_num, 2).value
    txn_date = coerce_date(raw_date)
    if txn_date is None:
        if last_seen_date is None:
            # Row has data but no date and no prior date — skip
            return None, last_seen_date, [
                f"Row {row_num}: no date and no prior date to forward-fill"
            ]
        txn_date = last_seen_date
    else:
        last_seen_date = txn_date

    # Required fields
    supplier = coerce_str(sheet.cell(row_num, 3).value)
    weight_kg = coerce_float(sheet.cell(row_num, 7).value)

    if supplier is None and weight_kg is None:
        # Blank-ish continuation row that snuck through — skip silently
        return None, last_seen_date, []

    if weight_kg is None:
        warnings.append(f"Row {row_num}: missing weight_kg — row skipped")
        return None, last_seen_date, warnings
    if not (WEIGHT_KG_MIN < weight_kg < WEIGHT_KG_MAX):
        warnings.append(
            f"Row {row_num}: weight {weight_kg} outside plausible range "
            f"({WEIGHT_KG_MIN}-{WEIGHT_KG_MAX})"
        )

    if supplier is None:
        warnings.append(f"Row {row_num}: missing supplier")

    # Block_loc — Blackwood regex check; tolerate operator's PILE_LOC continuation rows
    block_loc = coerce_str(sheet.cell(row_num, 5).value)
    if block_loc and not BLOCK_LOC_REGEX.match(block_loc):
        warnings.append(
            f"Row {row_num}: block_loc '{block_loc}' does not match "
            f"Blackwood format (e.g. A-1A, D-20D, PCA-15A)"
        )

    truck_plate = coerce_str(sheet.cell(row_num, 6).value)
    sacks = coerce_int(sheet.cell(row_num, 8).value)
    remarks = coerce_str(sheet.cell(row_num, 16).value)

    # Operator batch label and translation
    operator_batch_label = coerce_str(sheet.cell(row_num, 4).value)
    batch_code, batch_warnings = translate_batch_code(
        operator_batch_label, remarks, txn_date
    )
    warnings.extend(batch_warnings)

    # Lab metrics
    lab_results: dict[str, float | None] = {}
    for col, short_key in [
        (9, "mc"), (10, "grit"), (11, "bd_astm"), (12, "bd_jis"),
        (13, "vm"), (14, "ash"), (15, "fc"),
    ]:
        val = coerce_float(sheet.cell(row_num, col).value)
        lab_results[short_key] = val
        if val is not None and short_key in LAB_PLAUSIBILITY:
            msg, check = LAB_PLAUSIBILITY[short_key]
            if not check(val):
                warnings.append(f"Row {row_num}: {msg} ({short_key}={val})")

    # Weight deduction (ASH / MC / wet) annotated in the remark — additive,
    # display-only fields; weight_kg stays the deducted NET (see DEDUCTIONS_DESIGN.md).
    true_weight_kg, deduction_note, ded_warnings = detect_deduction(remarks, weight_kg)
    if ded_warnings:
        warnings.extend(f"Row {row_num}: {w}" for w in ded_warnings)

    # Confidence: start at 1.0, subtract 0.10 per warning (lighter than v1 since
    # batch-code translation always adds a warning), floor at 0.0
    confidence = max(0.0, 1.0 - 0.10 * len(warnings))

    row_dict = {
        "transaction_date": txn_date,
        "supplier": supplier,
        "batch_code": batch_code,
        "operator_batch_label": operator_batch_label,
        "block_loc": block_loc,
        "truck_plate": truck_plate,
        "sacks": sacks,
        "weight_kg": weight_kg,
        "cost_basis": None,  # operator file has no price columns
        "remarks": remarks,
        "lab_results": lab_results if any(v is not None for v in lab_results.values()) else None,
        # Deduction annotation (NULL on ordinary rows — every row carries both keys):
        "true_weight_kg": true_weight_kg,
        "deduction_note": deduction_note,
        "warnings": warnings,
        "confidence": round(confidence, 3),
        "_source_row": row_num,
    }
    return row_dict, last_seen_date, []


# ---------------------------------------------------------------------------
# Wet "recovery" sub-rows (see DEDUCTIONS_DESIGN.md Decision 8)
# ---------------------------------------------------------------------------
# A recovery sub-row is a continuation row directly under a full delivery (the
# "mother") that carries its OWN weight + sacks + MC but NO truck / batch label /
# block / supplier / date of its own. Historically these were dropped (silently,
# or flagged MALFORMED for lacking a batch_code) — the D-20D leak. We now emit
# each one as its OWN delivery row that INHERITS the mother's truck_plate,
# block_loc, supplier, batch_code (+ operator_batch_label), transaction_date and
# cost_basis, while keeping its own weight_kg / sacks / lab_results, plus its own
# true_weight_kg + deduction_note.
#
# The detection + builder live in lib.deductions (shared with extract_gsheet.py).
# This file only owns the email-specific "did the row have its OWN date?" check,
# which random-accesses the date cell — so is_recovery_candidate is a thin wrapper
# over the shared sheet-agnostic predicate.
def is_recovery_candidate(sheet, row_num: int, row_dict: dict[str, Any]) -> bool:
    """Email-path wrapper: compute has_own_date from the RAW col-2 date cell
    (extract_row forward-fills the date, so the row_dict value is non-null even on a
    recovery row) and delegate to the shared is_recovery_row_dict."""
    has_own_date = coerce_date(sheet.cell(row_num, 2).value) is not None
    return is_recovery_row_dict(row_dict, has_own_date=has_own_date)


def extract_sheet(sheet) -> tuple[list[dict[str, Any]], list[str]]:
    """Extract all data rows from one sheet. Returns (rows, sheet-level warnings)."""
    sheet_warnings: list[str] = []
    header_row = find_header_row(sheet)
    if header_row is None:
        sheet_warnings.append(
            f"Sheet '{sheet.title}': no recognizable header row found in first 15 rows"
        )
        return [], sheet_warnings

    data_start = first_data_row_below(sheet, header_row)
    rows: list[dict[str, Any]] = []
    last_seen_date: str | None = None
    # The most recent successfully-emitted MAIN (non-recovery) delivery row that a
    # recovery sub-row can inherit identity from.
    last_mother: dict[str, Any] | None = None

    for r in range(data_start, sheet.max_row + 1):
        row_dict, last_seen_date, extra = extract_row(sheet, r, last_seen_date)
        sheet_warnings.extend(extra)
        if row_dict is None:
            continue

        if is_recovery_candidate(sheet, r, row_dict):
            if _is_inheritable_mother(last_mother):
                recovery = build_recovery_row(row_dict, last_mother)
                rows.append(recovery)
                # A recovery does NOT become the mother for a subsequent recovery —
                # keep inheriting identity from the original mother delivery.
            else:
                # No mother to inherit from — keep the row as-is (it will surface
                # in MALFORMED at classify time, which is the correct signal that
                # an orphan recovery row appeared with no preceding delivery).
                sheet_warnings.append(
                    f"Row {r}: recovery-shaped sub-row with no preceding mother "
                    f"delivery to inherit from — left unmapped"
                )
                rows.append(row_dict)
            continue

        rows.append(row_dict)
        # Only a real, batch-carrying delivery row becomes the inheritance source.
        if _is_inheritable_mother(row_dict):
            last_mother = row_dict

    return rows, sheet_warnings


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="Extract RC DELIVERIES rows from XLSX.")
    parser.add_argument("--file", required=True, help="Path to the .xlsx file")
    parser.add_argument(
        "--sheet",
        help="Specific sheet name to process. Default: active sheet (usually the latest month).",
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Process every sheet in the workbook (for full historical bootstrap).",
    )
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    # Select sheets to process
    if args.all_sheets:
        sheet_names = list(wb.sheetnames)
    elif args.sheet:
        if args.sheet not in wb.sheetnames:
            print(
                json.dumps(
                    {
                        "error": f"Sheet '{args.sheet}' not found",
                        "available": list(wb.sheetnames),
                    }
                ),
                file=sys.stderr,
            )
            return 2
        sheet_names = [args.sheet]
    else:
        # Default: active sheet (typically the most recent month)
        sheet_names = [wb.active.title]

    all_rows: list[dict[str, Any]] = []
    all_warnings: list[str] = []
    sheets_processed: list[str] = []

    for name in sheet_names:
        ws = wb[name]
        rows, warns = extract_sheet(ws)
        # Tag each row with which sheet it came from
        for row in rows:
            row["_source_sheet"] = name
        all_rows.extend(rows)
        all_warnings.extend(warns)
        sheets_processed.append(name)

    confidences = [r["confidence"] for r in all_rows]
    overall_confidence = (
        round(sum(confidences) / len(confidences), 3) if confidences else 0.0
    )

    # Surface batch labels we couldn't translate (batch_code == raw operator label)
    unmapped = sorted(set(
        r.get("operator_batch_label") or ""
        for r in all_rows
        if r.get("batch_code") == r.get("operator_batch_label")
        and r.get("operator_batch_label")
    ))

    output = {
        "filename": path.name,
        "sheets_processed": sheets_processed,
        "rows": all_rows,
        "summary": {
            "total_rows": len(all_rows),
            "extraction_warnings": all_warnings,
            "overall_confidence": overall_confidence,
            "unmapped_batches": unmapped,
        },
    }

    print(json.dumps(output, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
