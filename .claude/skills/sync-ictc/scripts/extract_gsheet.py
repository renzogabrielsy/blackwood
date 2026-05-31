#!/usr/bin/env python3
"""
GSHEET extractor — pulls RC IN + RC OUT from Renzo's link-shared Google Sheet
(exported as XLSX) and normalizes both into Blackwood-shaped JSON.

This is the deterministic half of the `gsheet-sync` employee. The Sheet is
maintained by Renzo's own hires (derived from his master file MINUS pricing),
making it an independent source of truth for RC IN (deliveries) and RC OUT.

Source layout (verified empirically 2026-05-30 against the live Sheet):

  TAB "RC IN"  — header on ROW 7, data rows 8..end (~962 populated of 2,985 max).
    A=STATE   B=WHSE   C=DATE   D=SUPPLIER   E=BLOCK(=batch_code)  F=BLOCK LOC
    G=TRK(=truck_plate) H=WT(=weight_kg) I=SKS(=sacks)
    J=MC K=GRIT L=ASTM(=bd_astm) M=JIS(=bd_jis) N=VM O=ASH P=FC
    Q=REMARKS   R..X = WTD* weighted products (weight*metric) — IGNORED.
    NOTE: column E already holds Blackwood-style full batch_codes
    (e.g. "MAY-26-BLK13", "OCT-23-BLK1"), unlike the operator email file
    which holds short labels like "B09". So translation is light here — we
    still emit primary + fallback codes to survive the INCONSISTENT month
    prefix conventions (JAN vs JANUARY, MARCH vs MAR, etc.).
    Early "2023 BACKLOG" rows are sparse (often only WT + MC) — handled.
    No price column → cost_basis is always null (priced by the email side).

  TAB "RC OUT" — real header on ROW 4 (rows 1-3 are blank/title), data 5..end
    (~1,909 populated of 2,602 max).
    A=DATE  B=BATCH(month label only, e.g. "MAY" → production_batch month)
    C=BLOCK(=batch_code, e.g. "MARCH-26-BLK19")  D=WT(=weight_kg)
    E=PLANT/ETC(=destination: MAIN | SUNDRY, plus rare typos)  F=REMARKS
    G=BLOCK LOC  H=(blank)  I=MC  J=MC WTD  K=DAY  L=BATCH(dup)
    Critical: the batch_code lives in column C, NOT column B.

Usage:
    python3 extract_gsheet.py --file /tmp/gsheet_sync/rc_gsheet.xlsx \
        --out-rc-in /tmp/gsheet_sync/rc_in_extract.json \
        --out-rc-out /tmp/gsheet_sync/rc_out_extract.json

    # or print a combined object to stdout (no --out-* flags):
    python3 extract_gsheet.py --file /tmp/gsheet_sync/rc_gsheet.xlsx

Output (each tab): { tab, source_rows, rows[], summary{...} }
Each RC IN row carries primary + fallback batch_codes so the classifier can
resolve against whatever prefix convention the DB actually used.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
          file=sys.stderr)
    sys.exit(3)


# ---------------------------------------------------------------------------
# Blackwood batch_code month-prefix conventions — INCONSISTENT in the DB.
# (see CLAUDE.md / MEMORY: JAN/FEB are 3-letter, MARCH/APRIL/MAY full-name, ...)
# We derive a *fallback* alternate-prefix variant for each batch_code so the
# classifier can try primary first, then the alternate spelling.
# ---------------------------------------------------------------------------
# Canonical pairs of equivalent month prefixes seen across DB history.
MONTH_PREFIX_ALIASES: dict[str, str] = {
    "JAN": "JANUARY", "JANUARY": "JAN",
    "FEB": "FEBRUARY", "FEBRUARY": "FEB",
    "MARCH": "MAR", "MAR": "MARCH",
    "APRIL": "APR", "APR": "APRIL",
    # MAY has no alias (already 3 letters / full)
    "JUNE": "JUN", "JUN": "JUNE",
    "JULY": "JUL", "JUL": "JULY",
    "AUG": "AUGUST", "AUGUST": "AUG",
    "SEPT": "SEPTEMBER", "SEP": "SEPTEMBER", "SEPTEMBER": "SEPT",
    "OCT": "OCTOBER", "OCTOBER": "OCT",
    "NOV": "NOVEMBER", "NOVEMBER": "NOV",
    "DEC": "DECEMBER", "DECEMBER": "DEC",
}

# A batch_code looks like "<MONTHPREFIX>-<YY>-<SUFFIX>" e.g. "MARCH-26-BLK19",
# "OCT-23-BLK1", "JAN-26-SUNDRY5", "FEB-26-FEED6", "AUG-23-TNK2".
BATCH_CODE_RE = re.compile(r"^([A-Z]+)-(\d{2})-(.+)$", re.IGNORECASE)

WEIGHT_KG_MIN = 0
WEIGHT_KG_MAX = 200_000  # RC OUT day-totals can be large

BLOCK_LOC_REGEX = re.compile(r"^(PCA|PCB|[A-DF])-\d{1,2}[A-D]$")

# Destination typo normalization for RC OUT (col E).
DEST_TYPO_FIX = {"MAN": "MAIN", "MIAN": "MAIN"}
VALID_DESTINATIONS = {"MAIN", "SUNDRY"}

# Lab plausibility (soft warnings only; matches extract_rc_deliveries.py)
LAB_PLAUSIBILITY: dict[str, tuple[str, Any]] = {
    "mc":  ("Moisture content unusually high",   lambda v: v < 20),
    "ash": ("Ash content unusually high",        lambda v: v < 10),
    "fc":  ("Fixed carbon unusually low",        lambda v: v > 60),
    "vm":  ("Volatile matter unusually high",    lambda v: v < 25),
    "grit": ("Grit value unusually high",        lambda v: v < 5),
    "bd_astm": ("BD ASTM out of expected range", lambda v: 0.2 < v < 1.0),
    "bd_jis":  ("BD JIS out of expected range",  lambda v: 0.2 < v < 1.0),
}


# ---------------------------------------------------------------------------
# Coercion helpers (shared shape with the email extractors)
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("₱", "").replace("$", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> int | None:
    f = coerce_float(value)
    return int(round(f)) if f is not None else None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def batch_code_fallbacks(batch_code: str | None) -> list[str]:
    """
    Given a batch_code, produce alternate-prefix spellings to try if the
    primary doesn't resolve against the DB. Handles the INCONSISTENT month
    prefix convention (e.g. 'MARCH-26-BLK6' <-> 'MAR-26-BLK6').
    Returns a de-duplicated list NOT including the primary.
    """
    if not batch_code:
        return []
    m = BATCH_CODE_RE.match(batch_code.strip())
    if not m:
        return []
    prefix, yy, suffix = m.group(1).upper(), m.group(2), m.group(3)
    fallbacks: list[str] = []
    alias = MONTH_PREFIX_ALIASES.get(prefix)
    if alias:
        fallbacks.append(f"{alias}-{yy}-{suffix}")
    # Also offer an upper-cased exact (harmless if identical) in case the
    # source had odd casing — keeps the classifier robust.
    upper = f"{prefix}-{yy}-{suffix}"
    if upper != batch_code.strip():
        fallbacks.append(upper)
    # De-dup preserving order
    seen: set[str] = set()
    out: list[str] = []
    for fb in fallbacks:
        if fb not in seen:
            seen.add(fb)
            out.append(fb)
    return out


# ---------------------------------------------------------------------------
# RC IN extraction
# ---------------------------------------------------------------------------
RC_IN_HEADER_ROW = 7
RC_IN_DATA_START = 8

# 1-based column index -> field
RC_IN_COLS = {
    1: "state", 2: "whse", 3: "transaction_date", 4: "supplier",
    5: "batch_code", 6: "block_loc", 7: "truck_plate", 8: "weight_kg",
    9: "sacks", 10: "lab_mc", 11: "lab_grit", 12: "lab_bd_astm",
    13: "lab_bd_jis", 14: "lab_vm", 15: "lab_ash", 16: "lab_fc", 17: "remarks",
}


def extract_rc_in(ws) -> dict[str, Any]:
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    source_rows = 0

    # read_only is forward-only — pull the full grid once.
    grid = list(ws.iter_rows(min_row=RC_IN_DATA_START, max_col=17, values_only=True))

    last_seen_date: str | None = None
    for offset, raw in enumerate(grid):
        rnum = RC_IN_DATA_START + offset

        def col(idx: int) -> Any:  # 1-based, like spreadsheet
            return raw[idx - 1] if idx - 1 < len(raw) else None

        txn_date = coerce_date(col(3))
        weight_kg = coerce_float(col(8))
        batch_code = coerce_str(col(5))
        supplier = coerce_str(col(4))

        # Skip fully-empty rows (the Sheet is padded with ~2,000 blank rows).
        if txn_date is None and weight_kg is None and batch_code is None and supplier is None:
            continue
        source_rows += 1

        row_warnings: list[str] = []

        # Forward-fill date for sparse continuation rows.
        if txn_date is None:
            if last_seen_date is None:
                row_warnings.append(f"Row {rnum}: no date and no prior date to fill")
            else:
                txn_date = last_seen_date
        else:
            last_seen_date = txn_date

        if weight_kg is None:
            row_warnings.append(f"Row {rnum}: missing weight_kg")
        elif not (WEIGHT_KG_MIN < weight_kg < WEIGHT_KG_MAX):
            row_warnings.append(f"Row {rnum}: weight {weight_kg} out of plausible range")

        block_loc = coerce_str(col(6))
        if block_loc and not BLOCK_LOC_REGEX.match(block_loc.upper()):
            row_warnings.append(f"Row {rnum}: block_loc '{block_loc}' off-format")

        truck_plate = coerce_str(col(7))
        sacks = coerce_int(col(9))
        remarks = coerce_str(col(17))

        # Lab metrics J..P
        lab_results: dict[str, float | None] = {}
        for cidx, key in [(10, "mc"), (11, "grit"), (12, "bd_astm"),
                          (13, "bd_jis"), (14, "vm"), (15, "ash"), (16, "fc")]:
            v = coerce_float(col(cidx))
            lab_results[key] = v
            if v is not None and key in LAB_PLAUSIBILITY:
                msg, check = LAB_PLAUSIBILITY[key]
                if not check(v):
                    row_warnings.append(f"Row {rnum}: {msg} ({key}={v})")

        confidence = max(0.0, 1.0 - 0.10 * len(row_warnings))

        rows.append({
            "transaction_date": txn_date,
            "supplier": supplier,
            "batch_code_primary": batch_code,
            "batch_code_fallbacks": batch_code_fallbacks(batch_code),
            "block_loc": block_loc,
            "truck_plate": truck_plate,
            "sacks": sacks,
            "weight_kg": weight_kg,
            "cost_basis": None,  # Sheet has no price column
            "remarks": remarks,
            "lab_results": lab_results if any(v is not None for v in lab_results.values()) else None,
            "warnings": row_warnings,
            "confidence": round(confidence, 3),
            "_source_row": rnum,
            "_source_tab": "RC IN",
        })
        warnings.extend(row_warnings)

    confidences = [r["confidence"] for r in rows]
    return {
        "tab": "RC IN",
        "source_rows": source_rows,
        "rows": rows,
        "summary": {
            "total_rows": len(rows),
            "overall_confidence": round(sum(confidences) / len(confidences), 3) if confidences else 0.0,
            "warnings_count": len(warnings),
            "extraction_warnings": warnings[:50],  # cap to keep JSON sane
        },
    }


# ---------------------------------------------------------------------------
# RC OUT extraction
# ---------------------------------------------------------------------------
RC_OUT_HEADER_ROW = 4
RC_OUT_DATA_START = 5

# 1-based column index -> field
RC_OUT_COLS = {
    1: "transaction_date", 2: "production_batch", 3: "batch_code",
    4: "weight_kg", 5: "destination", 6: "remarks", 7: "block_loc",
    9: "lab_mc", 10: "lab_mc_wtd", 11: "day",
}


def extract_rc_out(ws) -> dict[str, Any]:
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    source_rows = 0

    grid = list(ws.iter_rows(min_row=RC_OUT_DATA_START, max_col=12, values_only=True))

    for offset, raw in enumerate(grid):
        rnum = RC_OUT_DATA_START + offset

        def col(idx: int) -> Any:
            return raw[idx - 1] if idx - 1 < len(raw) else None

        txn_date = coerce_date(col(1))
        batch_code = coerce_str(col(3))   # batch_code lives in col C (BLOCK)
        weight_kg = coerce_float(col(4))

        # Skip blank / stray rows (e.g. a lone "J=0" weighted-avg artifact).
        if txn_date is None and batch_code is None and weight_kg is None:
            continue
        source_rows += 1

        row_warnings: list[str] = []
        if txn_date is None:
            row_warnings.append(f"Row {rnum}: missing transaction_date")
        if weight_kg is None:
            row_warnings.append(f"Row {rnum}: missing weight_kg")
        elif not (WEIGHT_KG_MIN < weight_kg < WEIGHT_KG_MAX):
            row_warnings.append(f"Row {rnum}: weight {weight_kg} out of plausible range")

        # Destination — normalize obvious typos, default MAIN.
        dest_raw = coerce_str(col(5))
        destination = "MAIN"
        if dest_raw:
            up = dest_raw.upper()
            up = DEST_TYPO_FIX.get(up, up)
            destination = up
            if up not in VALID_DESTINATIONS:
                row_warnings.append(f"Row {rnum}: unrecognized destination '{dest_raw}' (kept as-is)")

        production_batch = coerce_str(col(2))   # month label, e.g. "MAY"
        remarks = coerce_str(col(6))
        block_loc = coerce_str(col(7))

        confidence = max(0.0, 1.0 - 0.10 * len(row_warnings))

        rows.append({
            "transaction_date": txn_date,
            "batch_code_primary": batch_code,
            "batch_code_fallbacks": batch_code_fallbacks(batch_code),
            "production_batch": production_batch,
            "destination": destination,
            "weight_kg": weight_kg,
            "block_loc": block_loc,
            "remarks": remarks,
            "warnings": row_warnings,
            "confidence": round(confidence, 3),
            "_source_row": rnum,
            "_source_tab": "RC OUT",
        })
        warnings.extend(row_warnings)

    confidences = [r["confidence"] for r in rows]
    return {
        "tab": "RC OUT",
        "source_rows": source_rows,
        "rows": rows,
        "summary": {
            "total_rows": len(rows),
            "overall_confidence": round(sum(confidences) / len(confidences), 3) if confidences else 0.0,
            "warnings_count": len(warnings),
            "extraction_warnings": warnings[:50],
        },
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Extract RC IN + RC OUT from the Google Sheet XLSX.")
    ap.add_argument("--file", required=True, help="Path to the exported workbook XLSX")
    ap.add_argument("--out-rc-in", help="Write RC IN extract JSON here")
    ap.add_argument("--out-rc-out", help="Write RC OUT extract JSON here")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    for needed in ("RC IN", "RC OUT"):
        if needed not in wb.sheetnames:
            print(json.dumps({"error": f"Tab '{needed}' not found. Tabs: {wb.sheetnames}"}),
                  file=sys.stderr)
            return 2

    rc_in = extract_rc_in(wb["RC IN"])
    rc_out = extract_rc_out(wb["RC OUT"])

    if args.out_rc_in:
        Path(args.out_rc_in).write_text(json.dumps(rc_in, indent=2, default=str))
    if args.out_rc_out:
        Path(args.out_rc_out).write_text(json.dumps(rc_out, indent=2, default=str))

    # Compact summary to stdout for the calling agent.
    print(json.dumps({
        "ok": True,
        "rc_in": {**rc_in["summary"], "out": args.out_rc_in},
        "rc_out": {**rc_out["summary"], "out": args.out_rc_out},
    }, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
