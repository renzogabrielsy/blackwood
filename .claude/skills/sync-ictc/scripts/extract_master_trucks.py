#!/usr/bin/env python3
"""
MASTER TRUCKS sheet extractor — backfill script for truck_readings table.

Source: MASTER - ICTC INPUT FILE V1.xlsx, sheet "TRUCKS"
Structure: Two sections:

  SECTION A — Monthly summaries (rows 3-14): 2026 first-of-month per truck.
    Col 5  = date (first-of-month)
    Col 6  = AAV 6111 START KM
    Col 7  = AAV 6111 END KM
    Col 8  = AAV 6111 TTL KM
    Col 9  = AAV 6111 TTL FUEL
    Col 12 = AAV 6111 REMARKS
    Col 14 = KCA 378 START KM
    Col 15 = KCA 378 END KM
    Col 16 = KCA 378 TTL KM
    Col 17 = KCA 378 TTL FUEL
    Col 20 = KCA 378 REMARKS
    Col 22 = FORKLIFT START KM
    Col 23 = FORKLIFT END KM
    Col 24 = FORKLIFT TTL KM
    Col 25 = FORKLIFT TTL FUEL
    Col 28 = FORKLIFT REMARKS
    Skip rows where TTL KM == 0 (future months or truck not used).

  SECTION B — Daily detail (rows 18+): 2026 daily readings per truck.
    Col 1  = date
    Col 6  = AAV 6111 START
    Col 7  = AAV 6111 END
    Col 8  = AAV 6111 TTL DISTANCE
    Col 9  = AAV 6111 FUEL ISSUED
    Col 12 = AAV 6111 REMARKS
    Col 14 = KCA 378 START
    Col 15 = KCA 378 END
    Col 16 = KCA 378 TTL DISTANCE
    Col 17 = KCA 378 FUEL ISSUED
    Col 20 = KCA 378 REMARKS
    Col 22 = FORKLIFT START
    Col 23 = FORKLIFT END
    Col 24 = FORKLIFT TTL DISTANCE
    Col 25 = FORKLIFT FUEL ISSUED
    Col 28 = FORKLIFT REMARKS
    Skip rows where TTL KM == 0 (no trip that day).

FLAGGED conditions:
  - end_km < start_km (odometer regression — indicates data entry error)
  - Monthly TTL KM > 20,000 (anomalous — indicates misplaced cumulative km)
  - AAV February 2026: start_km=0 which causes TTL=156595.5 — known anomaly, FLAGGED

DB schema: UNIQUE(reading_date, plate_no).

Usage:
    python3 extract_master_trucks.py --file "path/to/MASTER.xlsx"
    python3 extract_master_trucks.py --file "..." --output /tmp/trucks_extract.json

Output:
    {
      "source_file": "...",
      "source_sheet": "TRUCKS",
      "rows": [ {truck_readings row}, ... ],
      "summary": {
        "total_rows": N,
        "monthly_rows": N,
        "daily_rows": N,
        "plates": {"AAV 6111": N, "KCA 378": N, "FORKLIFT": N},
        "flagged_rows": N,
        "date_range": "YYYY-MM-DD -> YYYY-MM-DD",
        "warnings": [...],
      }
    }
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)


# ---------------------------------------------------------------------------
# Column config for each truck in monthly section (col 5 = date) and daily (col 1 = date)
# ---------------------------------------------------------------------------
TRUCK_MONTHLY_COLS = {
    "AAV 6111": {
        "start":   6,
        "end":     7,
        "ttl":     8,
        "fuel":    9,
        "remarks": 12,
    },
    "KCA 378": {
        "start":   14,
        "end":     15,
        "ttl":     16,
        "fuel":    17,
        "remarks": 20,
    },
    "FORKLIFT": {
        "start":   22,
        "end":     23,
        "ttl":     24,
        "fuel":    25,
        "remarks": 28,
    },
}

# Daily section uses the same column indices for trucks, but col 1 = date (not col 5)
# The daily section starts at row 18 (row 16-17 = headers)
MONTHLY_DATE_COL  = 5
DAILY_DATE_COL    = 1

MONTHLY_DATA_START = 3
MONTHLY_DATA_END   = 14
DAILY_DATA_START   = 18

# Anomaly threshold: monthly TTL KM > this is likely a data entry error
MAX_REASONABLE_MONTHLY_KM = 20_000.0


# ---------------------------------------------------------------------------
# Coercion helpers
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def coerce_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    if isinstance(value, str) and "VALUE" in value:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return default
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Build one truck row
# ---------------------------------------------------------------------------
def build_truck_row(
    reading_date: str,
    plate_no: str,
    start_km: float | None,
    end_km: float | None,
    ttl_km: float | None,
    fuel_liters: float | None,
    remarks: str | None,
    source_row: int,
    section: str,
    warnings: list[str],
) -> dict[str, Any] | None:
    """
    Return a row dict or None if the row should be skipped.
    Skip when TTL KM == 0 (no usage that period).
    """
    effective_ttl = ttl_km if ttl_km is not None else 0.0

    if effective_ttl == 0:
        return None  # No usage — skip

    row_flags = []

    # Validate end >= start
    if start_km is not None and end_km is not None:
        if float(end_km) < float(start_km):
            row_flags.append(
                f"end_km ({end_km}) < start_km ({start_km}) — odometer regression"
            )

    # Validate: monthly TTL not anomalously large
    if section == "monthly" and effective_ttl > MAX_REASONABLE_MONTHLY_KM:
        row_flags.append(
            f"Monthly TTL KM = {effective_ttl:.1f} exceeds {MAX_REASONABLE_MONTHLY_KM:,.0f} km "
            f"(start_km={start_km}) — likely data entry error (e.g., cumulative KM used instead of monthly diff)"
        )

    # AAV Feb 2026 known anomaly: start=0 → huge TTL
    if plate_no == "AAV 6111" and section == "monthly" and start_km == 0.0 and effective_ttl > 10_000:
        if not any("start_km=0" in f for f in row_flags):
            row_flags.append(
                f"AAV Feb 2026: start_km=0, TTL={effective_ttl:.1f} km. "
                f"Likely start_km was left blank (should be ~154,695 from Jan end). "
                f"Monthly KM is overstated."
            )

    if row_flags:
        for flag in row_flags:
            warnings.append(f"TRUCKS {section} row {source_row} ({reading_date}) [{plate_no}]: {flag}")

    return {
        "reading_date": reading_date,
        "plate_no": plate_no,
        "start_km": float(start_km) if start_km is not None else 0.0,
        "end_km": float(end_km) if end_km is not None else 0.0,
        "ttl_km": effective_ttl,
        "fuel_liters": fuel_liters,
        "remarks": remarks,
        "_source_row": source_row,
        "_section": section,
        "_flagged": bool(row_flags),
        "_flags": row_flags,
    }


# ---------------------------------------------------------------------------
# Section A: Monthly summaries (rows 3-14)
# ---------------------------------------------------------------------------
def extract_monthly_section(ws, warnings: list[str]) -> list[dict[str, Any]]:
    rows = []
    for r in range(MONTHLY_DATA_START, MONTHLY_DATA_END + 1):
        raw_date = ws.cell(r, MONTHLY_DATE_COL).value
        reading_date = coerce_date(raw_date)
        if reading_date is None:
            continue

        for plate_no, cols in TRUCK_MONTHLY_COLS.items():
            start_km   = coerce_float(ws.cell(r, cols["start"]).value)
            end_km     = coerce_float(ws.cell(r, cols["end"]).value)
            ttl_km     = coerce_float(ws.cell(r, cols["ttl"]).value)
            fuel       = coerce_float(ws.cell(r, cols["fuel"]).value)
            remarks    = coerce_str(ws.cell(r, cols["remarks"]).value)

            row = build_truck_row(
                reading_date, plate_no, start_km, end_km, ttl_km, fuel,
                remarks, r, "monthly", warnings
            )
            if row:
                rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Section B: Daily detail (rows 18+)
# ---------------------------------------------------------------------------
def extract_daily_section(ws, warnings: list[str]) -> list[dict[str, Any]]:
    rows = []
    max_row = ws.max_row

    for r in range(DAILY_DATA_START, max_row + 1):
        raw_date = ws.cell(r, DAILY_DATE_COL).value
        reading_date = coerce_date(raw_date)
        if reading_date is None:
            continue

        # Daily section uses same column indices as monthly for truck columns
        for plate_no, cols in TRUCK_MONTHLY_COLS.items():
            start_km = coerce_float(ws.cell(r, cols["start"]).value)
            end_km   = coerce_float(ws.cell(r, cols["end"]).value)
            ttl_km   = coerce_float(ws.cell(r, cols["ttl"]).value)
            fuel     = coerce_float(ws.cell(r, cols["fuel"]).value)
            remarks  = coerce_str(ws.cell(r, cols["remarks"]).value)

            row = build_truck_row(
                reading_date, plate_no, start_km, end_km, ttl_km, fuel,
                remarks, r, "daily", warnings
            )
            if row:
                rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract truck_readings from MASTER TRUCKS sheet."
    )
    parser.add_argument("--file", required=True, help="Path to MASTER ICTC INPUT FILE V1.xlsx")
    parser.add_argument("--output", default=None, help="Output JSON path (default: print to stdout)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    if "TRUCKS" not in wb.sheetnames:
        print(json.dumps({"error": "TRUCKS sheet not found", "available": list(wb.sheetnames)}), file=sys.stderr)
        return 2

    ws = wb["TRUCKS"]
    warnings: list[str] = []

    monthly_rows = extract_monthly_section(ws, warnings)
    daily_rows   = extract_daily_section(ws, warnings)

    all_rows = monthly_rows + daily_rows

    # Stats
    dates = [r["reading_date"] for r in all_rows]
    date_range = f"{min(dates)} -> {max(dates)}" if dates else "N/A"

    plates: dict[str, int] = {}
    for r in all_rows:
        p = r["plate_no"]
        plates[p] = plates.get(p, 0) + 1

    flagged_rows = [r for r in all_rows if r.get("_flagged")]

    output = {
        "source_file": path.name,
        "source_sheet": "TRUCKS",
        "rows": all_rows,
        "summary": {
            "total_rows": len(all_rows),
            "monthly_rows": len(monthly_rows),
            "daily_rows": len(daily_rows),
            "flagged_rows": len(flagged_rows),
            "plates": plates,
            "date_range": date_range,
            "total_warnings": len(warnings),
            "warnings": warnings,
        },
    }

    json_output = json.dumps(output, indent=2, default=str)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json_output)
        print(json.dumps({
            "ok": True,
            "total_rows": len(all_rows),
            "monthly_rows": len(monthly_rows),
            "daily_rows": len(daily_rows),
            "plates": plates,
            "flagged_rows": len(flagged_rows),
            "warnings": len(warnings),
            "output": str(out_path),
        }))
    else:
        print(json_output)

    return 0


if __name__ == "__main__":
    sys.exit(main())
