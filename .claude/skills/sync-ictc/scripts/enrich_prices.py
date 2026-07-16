#!/usr/bin/env python3
"""
Enrich RC DELIVERIES extracted rows with prices from Czarina's
"RAW CHARCOAL PURCHASES -Daily" XLSX (sender: czarinaloumaximoictc@gmail.com).

The two files don't share a date key — Czarina records the "Date of Del.paid"
(payment date, typically 1 day after delivery). So we match on
(supplier_normalized, truck_plate_normalized, weight_kg) with date as a
tiebreaker only.

Usage:
    python3 enrich_prices.py \\
        --extract-json /tmp/.../extract_latest.json \\
        --prices-xlsx /tmp/.../RAW_CHARCOAL_PURCHASES.xlsx \\
        --sheet "May 2026" \\
        --output /tmp/.../extract_enriched.json

Output: same shape as extract_latest.json but with cost_basis filled in
where a price match was found, plus a price_match field per row showing
the matched Czarina row's (date, php_per_kg) for traceability.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}), file=sys.stderr)
    sys.exit(3)


# ---------------------------------------------------------------------------
# Czarina file column layout (verified 2026-05-27)
# Header rows: R3 (top) + R4 (sub). Data starts R5.
# ---------------------------------------------------------------------------
CZARINA_COLS = {
    1: "date",          # "Date of Del.paid"
    2: "supplier",      # "Supplier's Name"
    3: "truck_plate",   # "Truck PLATE #"
    4: "piling_block",
    5: "sacks",
    6: "gross_weight",
    7: "mc_ash_less",
    8: "net_weight",    # "NET (#kilos)"
    9: "php_per_kg",    # "PHP/kg."
    10: "total_amount",
}


def norm_supplier(s):
    """Strip whitespace, lowercase, remove common prefixes."""
    if s is None:
        return None
    s = str(s).strip().lower()
    # Drop common name prefixes: "M. ", "B. ", "E. ", etc.
    s = re.sub(r"^[a-z]\.\s+", "", s)
    return s if s else None


def norm_truck(s):
    """Strip all whitespace and special chars, uppercase."""
    if s is None:
        return None
    return re.sub(r"[\s\-_]", "", str(s)).upper() or None


def norm_weight(v):
    if v is None:
        return None
    try:
        return round(float(v), 0)  # whole kg precision for matching
    except (TypeError, ValueError):
        return None


def coerce_date_str(value):
    """Return YYYY-MM-DD or None."""
    if value is None or value == "" or value == "-":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s or s == "-":
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def date_diff_days(a, b):
    """Absolute difference in days. Returns infinity if either is None."""
    if a is None or b is None:
        return float("inf")
    try:
        da = datetime.strptime(a, "%Y-%m-%d").date()
        db = datetime.strptime(b, "%Y-%m-%d").date()
        return abs((da - db).days)
    except (ValueError, TypeError):
        return float("inf")


def load_czarina_rows(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Load all data rows from the specified Czarina sheet."""
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        sys.stderr.write(
            json.dumps(
                {
                    "error": f"Sheet '{sheet_name}' not found",
                    "available": list(wb.sheetnames),
                }
            )
            + "\n"
        )
        sys.exit(2)
    ws = wb[sheet_name]

    rows = []
    last_seen_date = None
    # Data starts at R5 (R3-4 are headers)
    for r in range(5, ws.max_row + 1):
        supplier = ws.cell(r, 2).value
        if supplier is None or str(supplier).strip() == "":
            continue
        if str(supplier).strip().lower() in ("average", "total", "sum"):
            continue

        raw_date = ws.cell(r, 1).value
        date_str = coerce_date_str(raw_date)
        if date_str is None:
            date_str = last_seen_date
        else:
            last_seen_date = date_str

        net_weight = norm_weight(ws.cell(r, 8).value)
        if net_weight is None or net_weight <= 0:
            continue

        php = ws.cell(r, 9).value
        if php is None:
            continue
        try:
            php_val = float(php)
        except (TypeError, ValueError):
            continue

        rows.append(
            {
                "_source_row": r,
                "date": date_str,
                "supplier_raw": str(supplier).strip(),
                "supplier_norm": norm_supplier(supplier),
                "truck_raw": str(ws.cell(r, 3).value or "").strip(),
                "truck_norm": norm_truck(ws.cell(r, 3).value),
                "net_weight": net_weight,
                "php_per_kg": php_val,
            }
        )
    return rows


def build_price_index(czarina_rows):
    """Index by (supplier_norm, truck_norm, net_weight) -> list of rows."""
    idx = {}
    for r in czarina_rows:
        key = (r["supplier_norm"], r["truck_norm"], r["net_weight"])
        idx.setdefault(key, []).append(r)
    return idx


def match_price(extracted_row, price_index, max_date_drift_days=7):
    """
    Find best matching Czarina row for an extracted RC DELIVERIES row.
    Returns (php_per_kg, czarina_row_meta) or (None, None) if no match.

    Match logic:
      1. Build key from (norm_supplier, norm_truck, weight_kg)
      2. If no candidates, return None
      3. If 1 candidate, use it (warn if date drift > max_date_drift_days)
      4. If >1 candidates, pick the one with smallest date drift from extracted_row.transaction_date
    """
    key = (
        norm_supplier(extracted_row.get("supplier")),
        norm_truck(extracted_row.get("truck_plate")),
        norm_weight(extracted_row.get("weight_kg")),
    )
    candidates = price_index.get(key, [])
    if not candidates:
        return None, None

    if len(candidates) == 1:
        c = candidates[0]
        return c["php_per_kg"], c

    # Multiple candidates — pick the closest in date
    ex_date = extracted_row.get("transaction_date")
    candidates_sorted = sorted(candidates, key=lambda c: date_diff_days(ex_date, c["date"]))
    best = candidates_sorted[0]
    return best["php_per_kg"], best


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich RC DELIVERIES rows with prices from Czarina's file.")
    parser.add_argument("--extract-json", required=True, help="Output of extract_rc_deliveries.py")
    parser.add_argument("--prices-xlsx", required=True, help="Czarina's RAW CHARCOAL PURCHASES xlsx")
    parser.add_argument("--sheet", required=True, help="Sheet name in prices xlsx, e.g. 'May 2026'")
    parser.add_argument("--output", required=True, help="Output enriched JSON path")
    args = parser.parse_args()

    extract_path = Path(args.extract_json)
    prices_path = Path(args.prices_xlsx)
    output_path = Path(args.output)

    if not extract_path.exists():
        print(json.dumps({"error": f"Extract file not found: {extract_path}"}), file=sys.stderr)
        return 1
    if not prices_path.exists():
        print(json.dumps({"error": f"Prices file not found: {prices_path}"}), file=sys.stderr)
        return 1

    extracted = json.loads(extract_path.read_text())
    czarina_rows = load_czarina_rows(prices_path, args.sheet)
    price_index = build_price_index(czarina_rows)

    matched = 0
    unmatched = 0
    unmatched_rows = []

    for row in extracted["rows"]:
        php, meta = match_price(row, price_index)
        if php is not None:
            row["cost_basis"] = php
            row["price_match"] = {
                "czarina_row": meta["_source_row"],
                "czarina_date": meta["date"],
                "php_per_kg": php,
            }
            matched += 1
        else:
            row["price_match"] = None
            unmatched += 1
            unmatched_rows.append({
                "source_row": row.get("_source_row"),
                "date": row.get("transaction_date"),
                "supplier": row.get("supplier"),
                "truck_plate": row.get("truck_plate"),
                "weight_kg": row.get("weight_kg"),
            })

    extracted.setdefault("summary", {})["price_enrichment"] = {
        "czarina_file": str(prices_path),
        "czarina_sheet": args.sheet,
        "czarina_rows_loaded": len(czarina_rows),
        "matched_count": matched,
        "unmatched_count": unmatched,
        "unmatched_rows": unmatched_rows,
    }

    output_path.write_text(json.dumps(extracted, indent=2, default=str))

    print(f"Czarina rows loaded:    {len(czarina_rows)}")
    print(f"Extracted rows matched: {matched}")
    print(f"Extracted rows unmatched: {unmatched}")
    if unmatched_rows:
        print(f"\n=== Unmatched (need manual price entry) ===")
        for ur in unmatched_rows[:20]:
            print(f"  R{ur['source_row']:02} {ur['date']} | {ur['supplier']:12} | {ur['truck_plate']:10} | {ur['weight_kg']:>7}kg")
    return 0


if __name__ == "__main__":
    sys.exit(main())
