#!/usr/bin/env python3
"""
PROPOSED DAILY REPORT extractor — one sheet per day, multiple block sections per sheet.

File structure (verified 2026-05-27 with real operator file):
- One sheet per day (e.g., "MAY 26"). Sheet name = abbreviated month + day, year derived from filename.
- Each sheet contains 1-N block sections, each ~7 rows.
- Block section anatomy:
    Row 0: WHSE # | <block_loc> | ... | STRT. BAL  | <number>  | DONE FEEDING (sometimes)
    Row 1: BLOCK DATE | <YYYY-MM-DD> | ... | DAY TOTAL | <number>  | <supplier sometimes>
    Row 2: BLOCK NO.  | # <N>        | ... | END BAL.  | <number>  | <supplier sometimes>
    Row 3: Gross weight | pallet1_gross | pallet2_gross | ... | REMARKS | <DONE/etc>
    Row 4: Pallet     | pallet1_sacks | pallet2_sacks | ...
    Row 5: Net        | pallet1_net   | pallet2_net   | ...
    Row 6: (blank separator)
- Footer rows like "3X50 = NONE / RC = NONE" should be skipped.

Each block section -> ONE rc_out row with weight_kg = DAY TOTAL.

Usage:
    python3 extract_proposed_daily.py --file path/to/PROPOSED.xlsx --year 2026
    python3 extract_proposed_daily.py --file path/to/PROPOSED.xlsx --year 2026 --sheet "MAY 26"
    python3 extract_proposed_daily.py --file path/to/PROPOSED.xlsx --year 2026 --all-sheets

Output: JSON to stdout — { filename, sheets_processed, rows[], summary }
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}), file=sys.stderr)
    sys.exit(3)


# ---------------------------------------------------------------------------
# Blackwood batch_code prefix conventions (verified 2026-05-27 from DB)
# ---------------------------------------------------------------------------
PRIMARY_MONTH_PREFIX = {
    1: "JAN", 2: "FEB", 3: "MARCH", 4: "APRIL", 5: "MAY",
    6: "JUNE", 7: "JULY", 8: "AUG", 9: "SEPT",
    10: "OCT", 11: "NOV", 12: "DEC",
}

# Fallbacks: if primary doesn't exist in DB, try these alternates
FALLBACK_MONTH_PREFIX = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MAR", 4: "APR", 5: "MAY",
    6: "JUNE", 7: "JULY", 8: "AUGUST", 9: "SEPTEMBER",
    10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER",
}

# Canonical production_batch month names (BUG-005) — the FULL uppercase word for every
# month. This is NOT the batch_code prefix above (which keeps its own mixed convention);
# it is the rc_out.production_batch spelling, and it must stay identical to the TS port's
# workers/sync/src/lib/months.ts MONTH_NAMES.
MONTH_NAME_UPPER = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL", 5: "MAY", 6: "JUNE",
    7: "JULY", 8: "AUGUST", 9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER",
}

# Sheet name -> month/day map
MONTH_NAME_TO_NUM = {
    "JANUARY": 1, "JAN": 1, "FEBRUARY": 2, "FEB": 2,
    "MARCH": 3, "MAR": 3, "APRIL": 4, "APR": 4,
    "MAY": 5, "JUNE": 6, "JUN": 6, "JULY": 7, "JUL": 7,
    "AUGUST": 8, "AUG": 8, "SEPTEMBER": 9, "SEPT": 9, "SEP": 9,
    "OCTOBER": 10, "OCT": 10, "NOVEMBER": 11, "NOV": 11,
    "DECEMBER": 12, "DEC": 12,
}

# SHEET_NAME_RE — the DAY-TAB name. L-048 (2026-09-03, both engines in lockstep — see
# PORTING_DECISIONS.md "Post-port business-rule changes"): this was
# r"^([A-Z]+)\s+(\d{1,2})\s*$" — a space, and nothing else, between the month word and
# the day. MC's SEPTEMBER 2026 workbook names its tabs "Aug. 29", "Sep. 1", "SEP. 2" (a
# PERIOD after the abbreviation), so all three failed to parse, the extractor returned
# ZERO rows from a workbook full of feedings, and rc_out silently stopped at 2026-08-28.
# Same shape as L-039 (Czarina's "Aug. 2026") and L-042 ("FEEDING # 1"): a worksheet name
# a HUMAN typed must be parsed tolerantly, never matched against one spelling.
# Accepted now: optional period after the month token, any/no whitespace around the
# separator, an optional trailing period, any case. Still rejected: a non-month word
# ("SUMMARY", "Sheet1"), an out-of-range day ("Aug 32", caught by date()), and a 3-4 digit
# tail ("JANUARY 2026", the RC MOVEMENT month-tab shape).
SHEET_NAME_RE = re.compile(r"^\s*([A-Za-z]+)\s*\.?\s*(\d{1,2})\s*\.?\s*$", re.IGNORECASE)
BLOCK_NO_RE = re.compile(r"^\s*#\s*(\d+)\s*$")

# FEED-section WHSE-label detector (2026-07-11 fix, both engines in lockstep — see
# PORTING_DECISIONS.md "Post-port business-rule changes"). Whole-word match on FEED or
# FEEDING, case-insensitive: catches "FEEDING AREA" (legacy template) AND "FOR FEEDING"
# (the operator's current template). Conservative — a real block label like "A-16D" or
# "16A NEAR PATHWAY" contains no FEED/FEEDING word and does not match.
FEED_LABEL_RE = re.compile(r"\bFEED(?:ING)?\b")

WEIGHT_KG_MIN = 0
WEIGHT_KG_MAX = 200_000  # daily totals can be large


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> int | None:
    f = coerce_float(value)
    return int(f) if f is not None else None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_block_no(value: Any) -> int | None:
    s = coerce_str(value)
    if s is None:
        return None
    m = BLOCK_NO_RE.match(s)
    if m:
        return int(m.group(1))
    # Tolerate just an integer
    try:
        return int(s)
    except ValueError:
        return None


def sheet_name_to_date(sheet_name: str, year: int) -> date | None:
    m = SHEET_NAME_RE.match(sheet_name)
    if not m:
        return None
    month_name = m.group(1).upper()
    day = int(m.group(2))
    month_num = MONTH_NAME_TO_NUM.get(month_name)
    if month_num is None:
        return None
    try:
        return date(year, month_num, day)
    except ValueError:
        return None


def derive_batch_codes(
    block_date_val: date | None,
    block_no: int | None,
    is_feed: bool,
) -> tuple[list[str], list[str]]:
    """
    Given the BLOCK DATE and BLOCK NO. from the report, derive candidate batch_codes.
    Returns (primary_candidates, fallback_candidates).

    Standard block: "{PREFIX}-{YY}-BLK{N}"
    FEED batch:     "{PREFIX}-{YY}-FEED{N}"

    Year (YY) comes from BLOCK DATE year (2-digit).
    Month prefix tries primary first, then fallback.
    """
    if block_date_val is None or block_no is None:
        return [], []
    month = block_date_val.month
    yy = f"{block_date_val.year % 100:02d}"
    kind = "FEED" if is_feed else "BLK"
    primary = PRIMARY_MONTH_PREFIX[month]
    fallback = FALLBACK_MONTH_PREFIX[month]
    primary_code = f"{primary}-{yy}-{kind}{block_no}"
    fallback_code = f"{fallback}-{yy}-{kind}{block_no}"
    if primary_code == fallback_code:
        return [primary_code], []
    return [primary_code], [fallback_code]


# ---------------------------------------------------------------------------
# Block section detection + extraction
# ---------------------------------------------------------------------------
def find_block_section_starts(ws) -> list[int]:
    """Return row numbers (1-based) where 'WHSE #' appears in col A — start of each block section."""
    starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "WHSE" in str(v).upper() and "#" in str(v):
            starts.append(r)
    return starts


def extract_block_section(
    ws,
    start_row: int,
    transaction_date: date,
) -> tuple[dict[str, Any] | None, list[str]]:
    """
    Extract one block section starting at start_row (the WHSE # row).
    Returns (row_dict, warnings).
    """
    warnings: list[str] = []

    # Anchor cells (col B for the label values)
    whse = coerce_str(ws.cell(start_row, 2).value)
    block_date_raw = ws.cell(start_row + 1, 2).value
    block_no_raw = ws.cell(start_row + 2, 2).value

    # Right-side stats (typically col L = 12)
    strt_bal = coerce_float(ws.cell(start_row, 12).value)
    day_total = coerce_float(ws.cell(start_row + 1, 12).value)
    end_bal = coerce_float(ws.cell(start_row + 2, 12).value)
    remarks = coerce_str(ws.cell(start_row + 3, 12).value)

    # Status (col M = 13, on the WHSE row); supplier on BLOCK NO. row
    status = coerce_str(ws.cell(start_row, 13).value)
    supplier = coerce_str(ws.cell(start_row + 2, 13).value)

    # Pallet rows (rows 3, 4, 5 below WHSE) — pallet data starts at col B (=2)
    # Stop at the first cell whose value is the literal text 'REMARKS' (right-side label intrusion)
    pallets_gross: list[float] = []
    pallets_count: list[int] = []
    pallets_net: list[float] = []

    for c in range(2, ws.max_column + 1):
        v = ws.cell(start_row + 3, c).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in {"REMARKS", "DONE", "DONE FEEDING", "FOR FEEDING", "MC AVERAGE:", ""}:
            break
        f = coerce_float(v)
        if f is not None:
            pallets_gross.append(f)
            pn = coerce_int(ws.cell(start_row + 4, c).value)
            net = coerce_float(ws.cell(start_row + 5, c).value)
            pallets_count.append(pn if pn is not None else 0)
            pallets_net.append(net if net is not None else 0.0)

    # Skip section if it looks like a footer or otherwise empty
    if whse is None or whse.strip() == "":
        return None, []

    # Validate the section
    if day_total is None:
        # If DAY TOTAL is missing, try summing nets
        if pallets_net:
            day_total = sum(pallets_net)
            warnings.append(f"DAY TOTAL missing; derived from net pallets sum = {day_total}")
        else:
            return None, [f"Block section at R{start_row} ({whse}): no DAY TOTAL and no pallet nets"]

    if not (WEIGHT_KG_MIN < day_total < WEIGHT_KG_MAX):
        warnings.append(f"DAY TOTAL {day_total} outside plausible range")

    # FEED vs standard block detection (2026-07-11 fix — was a "FEEDING AREA"-only
    # substring match that missed the newer "FOR FEEDING" section header; see
    # FEED_LABEL_RE above).
    is_feed = bool(whse and FEED_LABEL_RE.search(whse.upper()))

    block_date = coerce_date(block_date_raw)
    block_no = parse_block_no(block_no_raw)

    # Derive batch_code candidates
    primary_codes, fallback_codes = derive_batch_codes(block_date, block_no, is_feed)
    if not primary_codes:
        warnings.append(
            f"Could not derive batch_code (block_date={block_date}, block_no={block_no})"
        )

    # block_loc: for standard blocks, the WHSE # value IS the block_loc. For FEED, no block_loc.
    block_loc = None if is_feed else whse

    # Derive rc_out remarks: if status indicates closing, set 'CLOSED'
    closing_phrases = {"DONE", "DONE FEEDING", "CLOSED"}
    rc_remarks = None
    is_closing = False
    for candidate in (status, remarks):
        if candidate and candidate.strip().upper() in closing_phrases:
            is_closing = True
            break
    if is_closing:
        rc_remarks = "CLOSED"
    elif remarks and remarks.strip().upper() != "FOR FEEDING":
        # Preserve genuinely informational remarks; skip "FOR FEEDING" which is just status
        rc_remarks = remarks

    # production_batch convention: the FULL uppercase month name, for all 12 months
    # (e.g. "JULY" for a transaction in July). BUG-005 (2026-07-17): this used to emit
    # the %b 3-letter abbreviation with full-name overrides for May/June ONLY, while the
    # Google Sheet writer wrote full names — so rc_out.production_batch held BOTH "JUL"
    # and "JULY", splitting the RC Movement campaign picker into two phantom campaigns
    # ("JUL-2026" vs "JULY-2026"). The Sheet's full-name spelling is canonical.
    # Changed in lockstep with the TS port (workers/sync/src/lib/months.ts).
    production_batch = MONTH_NAME_UPPER[transaction_date.month]

    confidence = max(0.0, 1.0 - 0.10 * len(warnings))

    row_dict = {
        "transaction_date": transaction_date.isoformat(),
        "whse_label": whse,
        "block_loc": block_loc,
        "block_date": block_date.isoformat() if block_date else None,
        "block_no": block_no,
        "is_feed": is_feed,
        "batch_code_primary": primary_codes[0] if primary_codes else None,
        "batch_code_fallbacks": fallback_codes,
        "supplier": supplier,
        "strt_bal_kg": strt_bal,
        "day_total_kg": day_total,
        "end_bal_kg": end_bal,
        "weight_kg": day_total,  # what goes into rc_out
        "destination": "MAIN",
        "production_batch": production_batch,
        "remarks": rc_remarks,
        "operator_status": status,
        "operator_remarks_raw": remarks,
        "pallets_gross": pallets_gross,
        "pallets_count": pallets_count,
        "pallets_net": pallets_net,
        "pallet_count": len(pallets_gross),
        "is_closing": is_closing,
        "warnings": warnings,
        "confidence": round(confidence, 3),
        "_source_row": start_row,
    }
    return row_dict, warnings


def extract_sheet(ws, year: int) -> tuple[list[dict], list[str]]:
    """Extract all block sections from one sheet. Returns (rows, sheet_warnings)."""
    sheet_warnings: list[str] = []
    txn_date = sheet_name_to_date(ws.title, year)
    if txn_date is None:
        sheet_warnings.append(f"Sheet '{ws.title}': cannot parse date from sheet name")
        return [], sheet_warnings

    rows: list[dict] = []
    for start_r in find_block_section_starts(ws):
        row_dict, extra = extract_block_section(ws, start_r, txn_date)
        sheet_warnings.extend(extra)
        if row_dict is not None:
            row_dict["_source_sheet"] = ws.title
            rows.append(row_dict)

    return rows, sheet_warnings


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="Extract PROPOSED DAILY REPORT rows.")
    parser.add_argument("--file", required=True)
    parser.add_argument("--year", type=int, required=True,
                        help="Year for sheet-name date parsing (e.g., 2026)")
    parser.add_argument("--sheet", help="Specific sheet name (e.g., 'MAY 26'). Default: latest sheet.")
    parser.add_argument("--all-sheets", action="store_true",
                        help="Process every sheet (for catch-up/backfill).")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open: {e}"}), file=sys.stderr)
        return 3

    # Determine which sheets to process
    if args.all_sheets:
        sheet_names = list(wb.sheetnames)
    elif args.sheet:
        if args.sheet not in wb.sheetnames:
            print(json.dumps({
                "error": f"Sheet '{args.sheet}' not found",
                "available": list(wb.sheetnames),
            }), file=sys.stderr)
            return 2
        sheet_names = [args.sheet]
    else:
        # Default: latest sheet (last one, typically most recent day)
        sheet_names = [wb.sheetnames[-1]]

    all_rows: list[dict] = []
    all_warnings: list[str] = []
    sheets_processed: list[str] = []

    for name in sheet_names:
        ws = wb[name]
        rows, warns = extract_sheet(ws, args.year)
        all_rows.extend(rows)
        all_warnings.extend(warns)
        sheets_processed.append(name)

    confs = [r["confidence"] for r in all_rows]
    overall_conf = round(sum(confs) / len(confs), 3) if confs else 0.0

    output = {
        "filename": path.name,
        "year": args.year,
        "sheets_processed": sheets_processed,
        "rows": all_rows,
        "summary": {
            "total_rows": len(all_rows),
            "extraction_warnings": all_warnings,
            "overall_confidence": overall_conf,
            "feed_rows": sum(1 for r in all_rows if r.get("is_feed")),
            "closing_rows": sum(1 for r in all_rows if r.get("is_closing")),
            "total_kg": round(sum(r.get("day_total_kg") or 0 for r in all_rows), 2),
        },
    }
    print(json.dumps(output, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
