#!/usr/bin/env python3
"""
WASTE PRODUCTION REPORT extractor (Ivy's email) — one sheet per MONTH, one row per day
(sometimes two rows/day split by shift).

File structure (verified 2026-05-29 against the live email file):
- One sheet per month. Sheet title = "MONTHNAME YYYY" (NOTE: some have a leading space,
  e.g. " APRIL 2026" — tolerated).
- Header spans rows 2-4; data starts ~row 5.
- Each data row = one (date, shift) waste record.

Verified positional column map (§15.3 / §15.4 of PRODUCTION_DESIGN.md —
confirmed 8/8 value-for-value against MASTER on 2026-05-22 + 2026-05-23):

    Ivy header             SACKS col (DROPPED)   KLS value col   ->  schema field
    R.S. #1 DUST (RS 1A)   B                     C                   rs1a_kg
    RS 1B                  D                     E                   rs1b_kg
    FILTER                 F                     G                   bf_kg
    RS 2&3                 H                     I                   rs23_kg
    R.S. 5                 J                     K                   rs5_kg
    UNCOOKED/SHELL         L                     M                   trml1_kg
    STONES                 N                     O                   trml2_kg
    GRIT                   P                     Q                   grit_kg

    A{r} = the day's date
    R{r} = TOTAL WASTE reported (recon check only — NOT stored; DB generates the total)
    S{r} = buyer note (e.g. "PCG/BUNAWAN") -> remarks (informational)
    V{r} = shift label (MORNING SHIFT / EVENING SHIFT) — present only on dual-shift days

The SACKS columns (B/D/F/H/J/L/N/P) are intentionally DROPPED — the schema
(production_waste) has no sacks columns.

Row filtering:
- A{r} must parse as a date. If not, SKIP the row — this cleanly drops the bottom
  column-SUM footer row (all KLS columns summed, blank date) and trailing 0-stub rows.
- A row whose date's month differs from the sheet's month is a carryover (e.g.
  2026-04-30 in the MAY sheet) — still extracted with its TRUE date, plus a warning.
  Classification dedupes later, so this is safe.

Shift normalization (§15.6):
- MORNING SHIFT -> "M"
- EVENING SHIFT -> "E"
- absent V (pre-2026-05-25 single daily-total rows) -> "M" (locked default —
  consistent with downtime attaching to M)

Usage:
    python3 extract_waste_production.py --file path/to/WASTE.xlsx
    python3 extract_waste_production.py --file path/to/WASTE.xlsx --sheet "MAY 2026"
    python3 extract_waste_production.py --file path/to/WASTE.xlsx --all-sheets

Output: JSON to stdout — { filename, sheets_processed, waste[], summary }
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}), file=sys.stderr)
    sys.exit(3)


# ---------------------------------------------------------------------------
# Month-name conventions
# ---------------------------------------------------------------------------
MONTH_NAME_TO_NUM = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5, "JUNE": 6,
    "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}

# production_batch = full English month name UPPERCASE derived from each row's actual date.
NUM_TO_MONTH_NAME = {v: k for k, v in MONTH_NAME_TO_NUM.items()}

# Sheet title like "MAY 2026" or " APRIL 2026" (leading/trailing whitespace tolerated).
SHEET_NAME_RE = re.compile(r"^\s*([A-Za-z]+)\s+(\d{4})\s*$")

# ---------------------------------------------------------------------------
# Column indices (1-based) — KLS value columns only (SACKS cols intentionally dropped)
# ---------------------------------------------------------------------------
COL_DATE   = 1   # A
COL_RS1A   = 3   # C
COL_RS1B   = 5   # E
COL_BF     = 7   # G   (Ivy header: FILTER)
COL_RS23   = 9   # I
COL_RS5    = 11  # K
COL_TRML1  = 13  # M   (Ivy header: UNCOOKED/SHELL)
COL_TRML2  = 15  # O   (Ivy header: STONES)
COL_GRIT   = 17  # Q
COL_TTL    = 18  # R   TOTAL WASTE reported (recon only)
COL_BUYER  = 19  # S   buyer note -> remarks
COL_SHIFT  = 22  # V   shift label (dual-shift days only)

DATA_START_ROW = 5  # Rows 1-4 = title + header

RECON_TOLERANCE_KG = 1.0


# ---------------------------------------------------------------------------
# Coercion helpers (mirrors extract_proposed_daily.py)
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if "VALUE" in s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if "VALUE" in value:
            return None
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def normalize_shift(value: Any) -> str:
    """MORNING SHIFT -> M, EVENING SHIFT -> E. Absent/unknown -> M (locked default)."""
    s = coerce_str(value)
    if s is None:
        return "M"
    up = s.upper()
    if "MORNING" in up:
        return "M"
    if "EVENING" in up:
        return "E"
    # Defensive: MC's label is NIGHT SHIFT but that's the same physical 2nd shift -> E.
    if "NIGHT" in up:
        return "E"
    return "M"


def sheet_name_to_month(sheet_name: str) -> tuple[int | None, str | None]:
    """Return (month_num, normalized_title) parsed from the sheet title, or (None, None)."""
    m = SHEET_NAME_RE.match(sheet_name)
    if not m:
        return None, None
    month_name = m.group(1).upper()
    month_num = MONTH_NAME_TO_NUM.get(month_name)
    return month_num, sheet_name.strip()


# ---------------------------------------------------------------------------
# Sheet extractor
# ---------------------------------------------------------------------------
def extract_sheet(ws, warnings: list[str]) -> list[dict[str, Any]]:
    """Extract all (date, shift) waste rows from one month-sheet."""
    sheet_month, sheet_title = sheet_name_to_month(ws.title)
    if sheet_month is None:
        warnings.append(f"Sheet '{ws.title}': cannot parse month from sheet title — sheet skipped")
        return []

    rows: list[dict[str, Any]] = []

    for r in range(DATA_START_ROW, ws.max_row + 1):
        txn_date = coerce_date(ws.cell(r, COL_DATE).value)

        # CRITICAL filter: a row without a valid date is either the bottom column-SUM
        # footer (blank A, all KLS cols summed) or a trailing 0-stub row. Skip cleanly.
        if txn_date is None:
            continue

        row_warnings: list[str] = []

        # Carryover row: date's month != the sheet's month (e.g. 2026-04-30 in MAY sheet).
        # Still extract with the TRUE date — classification dedupes later — but warn.
        if txn_date.month != sheet_month:
            row_warnings.append(
                f"Carryover date {txn_date.isoformat()} in sheet '{ws.title}' "
                f"(sheet month={sheet_month:02d}) — extracted with true date"
            )

        # 8 waste streams (KLS columns; missing -> 0.0, matching schema NOT NULL DEFAULT 0).
        streams = {
            "rs1a_kg":  coerce_float(ws.cell(r, COL_RS1A).value)  or 0.0,
            "rs1b_kg":  coerce_float(ws.cell(r, COL_RS1B).value)  or 0.0,
            "bf_kg":    coerce_float(ws.cell(r, COL_BF).value)    or 0.0,
            "rs23_kg":  coerce_float(ws.cell(r, COL_RS23).value)  or 0.0,
            "rs5_kg":   coerce_float(ws.cell(r, COL_RS5).value)   or 0.0,
            "trml1_kg": coerce_float(ws.cell(r, COL_TRML1).value) or 0.0,
            "trml2_kg": coerce_float(ws.cell(r, COL_TRML2).value) or 0.0,
            "grit_kg":  coerce_float(ws.cell(r, COL_GRIT).value)  or 0.0,
        }

        # Defensive: negative stream values shouldn't happen, but flag if they do.
        for field, val in streams.items():
            if val < 0:
                row_warnings.append(f"{field} is negative: {val}")

        shift = normalize_shift(ws.cell(r, COL_SHIFT).value)
        production_batch = NUM_TO_MONTH_NAME[txn_date.month]  # full month UPPERCASE
        ttl_reported = coerce_float(ws.cell(r, COL_TTL).value)
        remarks = coerce_str(ws.cell(r, COL_BUYER).value)

        summed = round(sum(streams.values()), 4)

        rows.append({
            "transaction_date": txn_date.isoformat(),
            "production_batch": production_batch,
            "shift": shift,
            **streams,
            "ttl_waste_kg_reported": ttl_reported,
            "remarks": remarks,
            "_source_sheet": ws.title,
            "_source_row": r,
            "_summed_kg": summed,
            "warnings": row_warnings,
        })

        for w in row_warnings:
            warnings.append(f"WASTE row {r} ({txn_date.isoformat()}): {w}")

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract production_waste rows from Ivy's WASTE PRODUCTION REPORT (one sheet per month)."
    )
    parser.add_argument("--file", required=True, help="Path to WASTE PRODUCTION REPORT xlsx")
    parser.add_argument("--sheet", help="Specific month sheet (e.g. 'MAY 2026'). Default: latest sheet.")
    parser.add_argument("--all-sheets", action="store_true",
                        help="Process every month sheet (for catch-up/backfill).")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    # Determine which sheets to process.
    if args.all_sheets:
        sheet_names = list(wb.sheetnames)
    elif args.sheet:
        if args.sheet not in wb.sheetnames:
            print(json.dumps({
                "error": f"Sheet '{args.sheet}' not found",
                "available": list(wb.sheetnames),
            }), file=sys.stderr)
            return 2
        sheet_names = [args.sheet]
    else:
        # Default: latest sheet = most recent month. Pick the parseable sheet with the
        # max (year, month); fall back to the last sheet if none parse.
        parsed = []
        for name in wb.sheetnames:
            month_num, _ = sheet_name_to_month(name)
            m = SHEET_NAME_RE.match(name)
            if month_num is not None and m:
                parsed.append((int(m.group(2)), month_num, name))
        if parsed:
            parsed.sort()
            sheet_names = [parsed[-1][2]]
        else:
            sheet_names = [wb.sheetnames[-1]]

    all_rows: list[dict] = []
    all_warnings: list[str] = []
    sheets_processed: list[str] = []

    for name in sheet_names:
        ws = wb[name]
        rows = extract_sheet(ws, all_warnings)
        all_rows.extend(rows)
        sheets_processed.append(name)

    # Reconciliation self-check: summed streams vs the reported TOTAL WASTE.
    # A mismatch means the column mapping is wrong.
    recon_mismatches = []
    for row in all_rows:
        reported = row.get("ttl_waste_kg_reported")
        summed = row["_summed_kg"]
        if reported is None:
            recon_mismatches.append({
                "date": row["transaction_date"],
                "shift": row["shift"],
                "summed": summed,
                "reported": None,
            })
        elif abs(summed - reported) > RECON_TOLERANCE_KG:
            recon_mismatches.append({
                "date": row["transaction_date"],
                "shift": row["shift"],
                "summed": summed,
                "reported": reported,
            })

    total_waste_kg = round(sum(r["_summed_kg"] for r in all_rows), 2)

    output = {
        "filename": path.name,
        "sheets_processed": sheets_processed,
        "waste": all_rows,
        "summary": {
            "sheets": len(sheets_processed),
            "waste_count": len(all_rows),
            "total_waste_kg": total_waste_kg,
            "recon_mismatches": recon_mismatches,
            "extraction_warnings": all_warnings,
        },
    }
    print(json.dumps(output, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
