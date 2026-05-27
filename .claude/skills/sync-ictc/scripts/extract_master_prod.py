#!/usr/bin/env python3
"""
MASTER PROD sheet extractor — backfill script for production_runs, production_downtime,
and production_waste tables.

Source: MASTER - ICTC INPUT FILE V1.xlsx, sheet "PROD"
Structure: 3 side-by-side sub-tables
  - Cols A-F (cols 1-6)  = PRODUCTION OUTPUT -> production_runs
  - Cols H-O (cols 8-15) = DOWNTIME          -> production_downtime
  - Cols Q-AJ (cols 17-36) = WASTE SUMMARY   -> production_waste

Data starts at row 3 (row 1 = section headers, row 2 = column headers).
Trailing rows 630-638 are blank (no #VALUE! errors observed — MASTER was clean at inspection).

Usage:
    python3 extract_master_prod.py --file "path/to/MASTER - ICTC INPUT FILE V1.xlsx"
    python3 extract_master_prod.py --file "..." --output /tmp/prod_extract.json

Output:
    {
      "source_file": "...",
      "source_sheet": "PROD",
      "runs": [ {production_run row}, ... ],
      "downtime": [ {production_downtime row}, ... ],
      "waste": [ {production_waste row}, ... ],
      "summary": {
        "runs_extracted": N,
        "downtime_extracted": N,
        "waste_extracted": N,
        "runs_date_range": "YYYY-MM-DD -> YYYY-MM-DD",
        "warnings": [...],
      }
    }
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}),
        file=sys.stderr,
    )
    sys.exit(3)


# ---------------------------------------------------------------------------
# Column indices (1-based) for each sub-table
# ---------------------------------------------------------------------------
# PRODUCTION OUTPUT (cols A-F)
COL_PROD_DATE    = 1
COL_PROD_BATCH   = 2
COL_PROD_GRADE   = 3
COL_PROD_SHIFT   = 4
COL_PROD_TTL_KG  = 5
COL_PROD_REMARKS = 6

# DOWNTIME (cols H-O = 8-15)
COL_DT_DATE      = 8
COL_DT_BATCH     = 9
COL_DT_SHIFT     = 10
COL_DT_SHIFT_HRS = 11
COL_DT_DT_HRS    = 12
COL_DT_DT_MINS   = 13
# col 14 = DT_TTL (computed), col 15 = TTL_HRS (computed) — not stored

# WASTE SUMMARY (cols Q-AJ = 17-36)
COL_W_DATE    = 17
COL_W_BATCH   = 18
COL_W_SHIFT   = 19
COL_W_RS1A    = 20
COL_W_SKS1    = 21   # TEXT — may be "3 bags"
COL_W_RS1B    = 22
COL_W_SKS2    = 23
COL_W_BF      = 24
COL_W_SKS3    = 25
COL_W_RS23    = 26
COL_W_SKS4    = 27
COL_W_RS5     = 28
COL_W_SKS5    = 29
COL_W_TRML1   = 30
COL_W_SKS6    = 31
COL_W_TRML2   = 32
COL_W_SKS7    = 33
COL_W_GRIT    = 34
# col 35 = TTL WASTE (computed), col 36 = PROD LOSS (computed) — not stored

VALID_GRADES = {"3X50", "6X50", "8X50", "2X6"}
VALID_SHIFTS = {"M", "E", "N"}

DATA_START_ROW = 3  # Row 1 = section headers, Row 2 = column headers


# ---------------------------------------------------------------------------
# Coercion helpers
# ---------------------------------------------------------------------------
def coerce_date(value: Any) -> str | None:
    """Return YYYY-MM-DD or None."""
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    s = str(value).strip()
    return s if s else None


def coerce_sacks_text(value: Any) -> str | None:
    """Keep sacks columns as raw text — may be int or string like '3 bags'."""
    if value is None:
        return None
    if isinstance(value, str) and "VALUE" in value:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Production runs extractor
# ---------------------------------------------------------------------------
def extract_runs(ws, warnings: list[str]) -> list[dict[str, Any]]:
    runs = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_date = ws.cell(r, COL_PROD_DATE).value
        batch = ws.cell(r, COL_PROD_BATCH).value
        grade = ws.cell(r, COL_PROD_GRADE).value
        shift = ws.cell(r, COL_PROD_SHIFT).value
        ttl_kg = ws.cell(r, COL_PROD_TTL_KG).value
        remarks = ws.cell(r, COL_PROD_REMARKS).value

        txn_date = coerce_date(raw_date)
        grade_s = coerce_str(grade)
        shift_s = coerce_str(shift)
        ttl_kg_f = coerce_float(ttl_kg)

        # Skip blank rows
        if txn_date is None and grade_s is None and ttl_kg_f is None:
            continue
        # Skip #VALUE! formula error rows
        if txn_date is None:
            if raw_date is not None:
                warnings.append(f"PROD row {r}: unreadable date '{raw_date}' — skipped")
            continue

        if ttl_kg_f is None:
            warnings.append(f"PROD row {r}: missing ttl_kg for date={txn_date} grade={grade_s} — skipped")
            continue

        # Validate grade
        row_warnings = []
        if grade_s not in VALID_GRADES:
            row_warnings.append(f"Unknown grade '{grade_s}'")

        # Validate shift
        if shift_s not in VALID_SHIFTS:
            row_warnings.append(f"Unknown shift '{shift_s}'")

        if ttl_kg_f < 0:
            row_warnings.append(f"Negative ttl_kg={ttl_kg_f}")

        if row_warnings:
            for w in row_warnings:
                warnings.append(f"PROD row {r} ({txn_date}): {w}")

        runs.append({
            "transaction_date": txn_date,
            "production_batch": coerce_str(batch) or "UNKNOWN",
            "grade": grade_s,
            "shift": shift_s,
            "ttl_kg": ttl_kg_f,
            "sacks_bags": None,   # Not in MASTER PROD sheet
            "remarks": coerce_str(remarks),
            "_source_row": r,
            "_warnings": row_warnings,
        })

    return runs


# ---------------------------------------------------------------------------
# Downtime extractor
# ---------------------------------------------------------------------------
def extract_downtime(ws, warnings: list[str]) -> list[dict[str, Any]]:
    downtime = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_date = ws.cell(r, COL_DT_DATE).value
        batch = ws.cell(r, COL_DT_BATCH).value
        shift = ws.cell(r, COL_DT_SHIFT).value
        shift_hrs = ws.cell(r, COL_DT_SHIFT_HRS).value
        dt_hrs = ws.cell(r, COL_DT_DT_HRS).value
        dt_mins = ws.cell(r, COL_DT_DT_MINS).value

        txn_date = coerce_date(raw_date)
        shift_s = coerce_str(shift)

        # Skip blank rows (downtime sub-table is sparse)
        if txn_date is None and shift_s is None:
            continue
        if txn_date is None:
            continue

        if shift_s is None:
            warnings.append(f"DOWNTIME row {r}: missing shift for date={txn_date} — skipped")
            continue

        shift_hrs_f = coerce_float(shift_hrs)
        dt_hrs_f = coerce_float(dt_hrs) or 0.0
        dt_mins_f = coerce_float(dt_mins) or 0.0

        if shift_hrs_f is None:
            warnings.append(f"DOWNTIME row {r} ({txn_date}): missing shift_hrs — skipped")
            continue

        row_warnings = []

        # Validate shift
        if shift_s not in VALID_SHIFTS:
            row_warnings.append(f"Unknown shift '{shift_s}'")

        # Validate dt_mins < 60
        if dt_mins_f >= 60:
            row_warnings.append(
                f"dt_mins={dt_mins_f} >= 60 (should be split into hours). "
                f"Computed dt_ttl_hrs = {dt_hrs_f + dt_mins_f/60:.3f}. "
                f"Row still inserted as-is — verify with operator."
            )

        # Validate total downtime <= shift_hrs
        dt_total = dt_hrs_f + dt_mins_f / 60.0
        if dt_total > shift_hrs_f:
            row_warnings.append(
                f"Downtime total ({dt_total:.3f} hrs) > shift_hrs ({shift_hrs_f})"
            )

        if row_warnings:
            for w in row_warnings:
                warnings.append(f"DOWNTIME row {r} ({txn_date}): {w}")

        downtime.append({
            "transaction_date": txn_date,
            "production_batch": coerce_str(batch) or "UNKNOWN",
            "shift": shift_s,
            "shift_hrs": shift_hrs_f,
            "dt_hrs": dt_hrs_f,
            "dt_mins": dt_mins_f,
            "dt_reason": None,  # Not in MASTER; will come from MC's daily email
            "_source_row": r,
            "_warnings": row_warnings,
        })

    return downtime


# ---------------------------------------------------------------------------
# Waste extractor
# ---------------------------------------------------------------------------
def extract_waste(ws, warnings: list[str]) -> list[dict[str, Any]]:
    waste = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_date = ws.cell(r, COL_W_DATE).value
        batch = ws.cell(r, COL_W_BATCH).value
        shift = ws.cell(r, COL_W_SHIFT).value

        txn_date = coerce_date(raw_date)
        shift_s = coerce_str(shift)

        # Skip blank rows
        if txn_date is None and shift_s is None:
            continue
        if txn_date is None:
            continue

        if shift_s is None:
            warnings.append(f"WASTE row {r}: missing shift for date={txn_date} — skipped")
            continue

        row_warnings = []

        # Validate shift
        if shift_s not in VALID_SHIFTS:
            row_warnings.append(f"Unknown shift '{shift_s}'")

        # Extract all 8 waste streams
        streams = {
            "rs1a_kg": coerce_float(ws.cell(r, COL_W_RS1A).value) or 0.0,
            "rs1a_sacks": coerce_sacks_text(ws.cell(r, COL_W_SKS1).value),
            "rs1b_kg": coerce_float(ws.cell(r, COL_W_RS1B).value) or 0.0,
            "rs1b_sacks": coerce_sacks_text(ws.cell(r, COL_W_SKS2).value),
            "bf_kg":   coerce_float(ws.cell(r, COL_W_BF).value) or 0.0,
            "bf_sacks":   coerce_sacks_text(ws.cell(r, COL_W_SKS3).value),
            "rs23_kg": coerce_float(ws.cell(r, COL_W_RS23).value) or 0.0,
            "rs23_sacks": coerce_sacks_text(ws.cell(r, COL_W_SKS4).value),
            "rs5_kg":  coerce_float(ws.cell(r, COL_W_RS5).value) or 0.0,
            "rs5_sacks":  coerce_sacks_text(ws.cell(r, COL_W_SKS5).value),
            "trml1_kg": coerce_float(ws.cell(r, COL_W_TRML1).value) or 0.0,
            "trml1_sacks": coerce_sacks_text(ws.cell(r, COL_W_SKS6).value),
            "trml2_kg": coerce_float(ws.cell(r, COL_W_TRML2).value) or 0.0,
            "trml2_sacks": coerce_sacks_text(ws.cell(r, COL_W_SKS7).value),
            "grit_kg": coerce_float(ws.cell(r, COL_W_GRIT).value) or 0.0,
        }

        # Validate: no negative stream kg
        for field, val in streams.items():
            if field.endswith("_kg") and isinstance(val, float) and val < 0:
                row_warnings.append(f"{field} is negative: {val}")

        if row_warnings:
            for w in row_warnings:
                warnings.append(f"WASTE row {r} ({txn_date}): {w}")

        waste.append({
            "transaction_date": txn_date,
            "production_batch": coerce_str(batch) or "UNKNOWN",
            "shift": shift_s,
            **streams,
            "remarks": None,
            "_source_row": r,
            "_warnings": row_warnings,
        })

    return waste


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract production_runs, production_downtime, production_waste from MASTER PROD sheet."
    )
    parser.add_argument("--file", required=True, help="Path to MASTER ICTC INPUT FILE V1.xlsx")
    parser.add_argument("--output", default=None, help="Output JSON path (default: print to stdout)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open XLSX: {e}"}), file=sys.stderr)
        return 3

    if "PROD" not in wb.sheetnames:
        print(json.dumps({"error": "PROD sheet not found", "available": list(wb.sheetnames)}), file=sys.stderr)
        return 2

    ws = wb["PROD"]
    warnings: list[str] = []

    runs = extract_runs(ws, warnings)
    downtime = extract_downtime(ws, warnings)
    waste = extract_waste(ws, warnings)

    # Date range for runs
    run_dates = [r["transaction_date"] for r in runs if r["transaction_date"]]
    date_range = f"{min(run_dates)} -> {max(run_dates)}" if run_dates else "N/A"

    output = {
        "source_file": path.name,
        "source_sheet": "PROD",
        "runs": runs,
        "downtime": downtime,
        "waste": waste,
        "summary": {
            "runs_extracted": len(runs),
            "downtime_extracted": len(downtime),
            "waste_extracted": len(waste),
            "runs_date_range": date_range,
            "total_warnings": len(warnings),
            "warnings": warnings,
        },
    }

    json_output = json.dumps(output, indent=2, default=str)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json_output)
        print(json.dumps({
            "ok": True,
            "runs": len(runs),
            "downtime": len(downtime),
            "waste": len(waste),
            "warnings": len(warnings),
            "output": str(out_path),
        }))
    else:
        print(json_output)

    return 0


if __name__ == "__main__":
    sys.exit(main())
