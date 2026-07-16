#!/usr/bin/env python3
"""
RAW CHARCOAL MOVEMENT extractor — daily totals + product breakdown.

This file is the **audit/reference** source — never writes to DB. Its main
role is to provide the daily "RAW CHARCOAL FED (KLS.)" total, which should
equal the sum of PROPOSED DAILY REPORT block sections for the same day.

File structure (verified 2026-05-27):
- One sheet per month (e.g., "MAY 2026")
- Header rows R1-R5 (title + multi-row header)
- Data rows R7+ : DATE | RAW CHARCOAL FED (KLS.) | DERAMI | OVER columns |
                  Re-Classified columns | Mixing | Blending columns | Production columns
- Sparse blank rows between dates

Output: per-date JSON with at minimum (date, raw_charcoal_fed_kls). Product
breakdown columns are optional in this v1 — captured if non-null but not yet
used for any cross-checks (will feed flecon_bag_movement / production tables
when those agents are built).

Usage:
    python3 extract_rc_movement.py --file path/to/RC_MOVEMENT.xlsx --sheet "MAY 2026"
    python3 extract_rc_movement.py --file path/to/RC_MOVEMENT.xlsx --all-sheets
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}), file=sys.stderr)
    sys.exit(3)


# Column 1 (A) = DATE, Column 2 (B) = RAW CHARCOAL FED (KLS.)
# Right-side columns are categorized product output; capture them as raw
# numeric values keyed by their header text. We attempt to read R3-R4 as
# the header pair.

WEIGHT_KG_MAX = 200_000


def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s or s == "-":
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned or cleaned == "-":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def build_column_headers(ws, header_rows: list[int]) -> dict[int, str]:
    """
    Combine multiple header rows into a single column-index -> string mapping.
    Joins non-empty cells across header rows with a space.
    """
    headers: dict[int, list[str]] = {}
    for r in header_rows:
        for c in range(1, ws.max_column + 1):
            v = coerce_str(ws.cell(r, c).value)
            if v:
                headers.setdefault(c, []).append(v)
    return {c: " ".join(parts).strip() for c, parts in headers.items()}


def extract_sheet(ws) -> tuple[list[dict], list[str]]:
    """Extract per-date rows from one RC MOVEMENT month sheet."""
    warnings: list[str] = []

    # Find header rows. Typical pattern: R3 + R4 (sometimes R3-R5)
    # We look for a row containing "DATE" in col A or B
    header_row = None
    for r in range(1, min(ws.max_row + 1, 10)):
        v = coerce_str(ws.cell(r, 1).value)
        if v and v.upper() == "DATE":
            header_row = r
            break
    if header_row is None:
        warnings.append(f"Sheet '{ws.title}': no DATE header row found")
        return [], warnings

    # Build column headers from R3+R4 (or header_row + header_row+1)
    headers = build_column_headers(ws, [header_row, header_row + 1, header_row + 2])

    rows: list[dict] = []
    # Data starts 3 rows below the DATE header (skip sub-header and unit row).
    # Stop at the first sentinel row that indicates a new section (e.g., "SUPPLIERS", "REMARKS:")
    # or any free-text in col A that isn't a date and isn't a known sub-header.
    SECTION_BREAK_TOKENS = {"SUPPLIERS", "REMARKS:", "REMARKS", "NOTES", "TOTAL", "TOTALS"}

    data_start = header_row + 3
    for r in range(data_start, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        # Check for section break — text in col A that isn't a date
        if isinstance(a_val, str):
            stripped = a_val.strip().upper()
            if stripped in SECTION_BREAK_TOKENS:
                break  # Stop at SUPPLIERS / REMARKS section

        d = coerce_date(a_val)
        if d is None:
            continue
        fed_kls = coerce_float(ws.cell(r, 2).value)
        if fed_kls is None:
            # Date present but no fed total — skip with note
            warnings.append(f"R{r}: date {d} present but RAW CHARCOAL FED is empty")
            continue
        if fed_kls < 0 or fed_kls > WEIGHT_KG_MAX:
            warnings.append(f"R{r}: implausible fed_kls={fed_kls}")

        # Capture other product columns for future use
        breakdown: dict[str, float] = {}
        for c in range(3, ws.max_column + 1):
            f = coerce_float(ws.cell(r, c).value)
            if f is None or f == 0:
                continue
            header = headers.get(c) or f"col_{c}"
            breakdown[header] = f

        rows.append({
            "transaction_date": d.isoformat(),
            "raw_charcoal_fed_kls": fed_kls,
            "product_breakdown": breakdown if breakdown else None,
            "_source_row": r,
            "_source_sheet": ws.title,
        })

    return rows, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract RC MOVEMENT daily totals (audit-only).")
    parser.add_argument("--file", required=True)
    parser.add_argument("--sheet", help="Sheet name e.g. 'MAY 2026'. Default: active sheet.")
    parser.add_argument("--all-sheets", action="store_true")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        return 1

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open: {e}"}), file=sys.stderr)
        return 3

    if args.all_sheets:
        sheet_names = list(wb.sheetnames)
    elif args.sheet:
        if args.sheet not in wb.sheetnames:
            print(json.dumps({
                "error": f"Sheet '{args.sheet}' not found",
                "available": list(wb.sheetnames),
            }), file=sys.stderr)
            return 2
        sheet_names = [args.sheet]
    else:
        sheet_names = [wb.active.title]

    all_rows: list[dict] = []
    all_warnings: list[str] = []
    for name in sheet_names:
        ws = wb[name]
        rows, warns = extract_sheet(ws)
        all_rows.extend(rows)
        all_warnings.extend(warns)

    # Build a date -> fed_kls lookup for convenient reconciliation.
    # A boundary date can appear on MORE THAN ONE month-tab (e.g. May 29 closes out
    # on the "MAY 2026" tab and opens on the "JUNE 2026" tab). SUM across tabs —
    # never overwrite — otherwise a cross-month date undercounts here, and the
    # DB-vs-RC-MOVEMENT duplication gate (L-019) reads the DB's correct cross-tab
    # total as a phantom excess and false-halts. (L-022)
    date_index: dict[str, float] = {}
    for r in all_rows:
        _d = r["transaction_date"]
        date_index[_d] = round(date_index.get(_d, 0.0) + r["raw_charcoal_fed_kls"], 2)

    output = {
        "filename": path.name,
        "sheets_processed": sheet_names,
        "rows": all_rows,
        "date_to_fed_kls": date_index,
        "summary": {
            "total_dates": len(all_rows),
            "total_fed_kls": round(sum(r["raw_charcoal_fed_kls"] for r in all_rows), 2),
            "date_range": [all_rows[0]["transaction_date"], all_rows[-1]["transaction_date"]]
                if all_rows else None,
            "extraction_warnings": all_warnings,
        },
    }
    print(json.dumps(output, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
