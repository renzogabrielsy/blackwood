#!/usr/bin/env python3
"""build_fixtures.py — synthesize edge-case workbooks with openpyxl, mirroring the
real workbook geometries so the SAME Python extractors read them. Each builder
also emits a hand-curated db_window snapshot so the oracle+TS classify against a
tiny, fully-controlled DB context that surfaces exactly one rule.

These synthetic cases are where the L-rule COVERAGE lives (the real-workbook
cases prove end-to-end parity but exercise mostly-NOOP tails). Every case is
registered in its type's manifest.json with a `covers` list and verified by
running the real extractor against it (build_oracle) — if the geometry is wrong,
the oracle build fails loudly rather than silently mis-reading.

Run:  python3 scripts/build_fixtures.py [--type <t>]
Then: python3 scripts/build_oracle.py   (regenerates oracles for the new cases)

Geometry references (read off the real workbooks 2026-07-04):
  flecon   : r1 title, r4 A=DATE B=PARTICULAR, r5 col-C+ bag signatures,
             r7 "Forwarded Balance" opening, r8+ data w/ date carry-forward.
  deliveries: r2 main header, r3 sub, r6+ data; fixed OPERATOR_COLUMNS 1..16.
  proposed : one sheet per day "MON DD"; 7-row block sections keyed on a col-A
             cell containing WHSE + '#'.
  gsheet   : RC IN header row 7 data 8+; RC OUT header row 4 data 5+.
  waste    : one sheet per month "MONTH YYYY"; header rows 2-4; data row 5+.
  rc_movement: one sheet per month; col-A 'DATE' header; data at header_row+3.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
WORKER = HERE.parent
FIXTURES = WORKER / "fixtures"


def _wb_dir(rt: str) -> Path:
    d = FIXTURES / rt / "workbooks"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _dbw_dir(rt: str) -> Path:
    d = FIXTURES / rt / "db_window"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _write_dbw(rt: str, case: str, bundle: dict) -> None:
    (_dbw_dir(rt) / f"{case}.json").write_text(json.dumps(bundle, indent=2, default=str))


# ─────────────────────────────────────────────────────────────────────────────
# FLECON — the reference synthetic (multi-row header, ambiguity, ZAMBOANGA typo,
# balance snapshot, multi-column row, unmapped column, day-set NOOP/CHANGED).
# ─────────────────────────────────────────────────────────────────────────────

def build_flecon() -> None:
    """One workbook exercising all flecon extraction+classify rules. bag_type_types
    registry is REORDERED vs column order to prove signature (not positional) mapping."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JANUARY 2026"
    # Title rows
    ws["E1"] = " FLECON BAG MOVEMENT"
    ws["E2"] = "YEAR 2025"  # deliberately stale — must be ignored
    # Row 4: DATE / PARTICULAR + section label
    ws["A4"] = "DATE"
    ws["B4"] = "PARTICULAR"
    ws["C4"] = "FLECON BAG"
    ws["H4"] = "Running Balance"
    # Row 5: per-column bag-type signatures (col C = 3 onward).
    #  C: KURARAY_590  D: UNUSABLE  E: FUTAMURA_550
    #  F/G: BOTH row-5 "FG" (ambiguous alone) — disambiguated by row 3.
    #  I: a column matching NO registry entry (UNMAPPED).
    ws["C5"] = "590 kls(Kuraray)"
    ws["D5"] = "Un-usable bag"
    ws["E5"] = "550 kls(FUTAMURA)"
    ws["F5"] = "FG"
    ws["G5"] = "FG"
    ws["I5"] = "MYSTERY BRAND XYZ"  # unmapped column
    # Row 3 disambiguators for F/G (combined signature = row3 + row5)
    ws["F3"] = "ALL BLACK"
    ws["G3"] = "BLACK SLING 6X50"
    # Row 7: Forwarded Balance (opening) — mix of populated + blank
    ws["B7"] = "Forwarded Balance"
    ws["C7"] = 20   # KURARAY_590 opening
    ws["E7"] = 507  # FUTAMURA opening
    # D7 blank -> omitted from opening dict

    # Data rows (row 8+). Date carry-forward: date only on first row of a day.
    r = 8
    # Day 2026-06-10: two movements, then a multi-column row (blend touching 2 types)
    ws.cell(r, 1, "2026-06-10"); ws.cell(r, 2, "BAGGED POWDER"); ws.cell(r, 3, -2); r += 1  # KURARAY_590 -2
    ws.cell(r, 2, "RS 1 ZAMBOANGA"); ws.cell(r, 4, -4); r += 1  # UNUSABLE col D -4 (marker text w/ qty)
    # multi-column single row: one date, two bag types (C and E)
    ws.cell(r, 2, "BLEND RECOUNT"); ws.cell(r, 3, 5); ws.cell(r, 5, 7); r += 1
    # a bare marker row (typo spelling) with NO quantity -> skipped_markers tally
    ws.cell(r, 2, "RS 1 ZAMBAONGA"); r += 1
    # month-section header row (resets carried_date)
    ws.cell(r, 1, "JULY"); r += 1
    # Day 2026-07-01 (after the month header — must NOT inherit June's date)
    ws.cell(r, 1, "2026-07-01"); ws.cell(r, 2, "BAGGED FG"); ws.cell(r, 6, -3); r += 1  # F col (ALL BLACK)
    ws.cell(r, 2, "BAGGED SLING"); ws.cell(r, 7, -1); r += 1  # G col (BLACK SLING)
    # balance-snapshot row: NO date, NO particular, has quantities
    ws.cell(r, 3, 18); ws.cell(r, 5, 514); r += 1

    _wb_dir("flecon")
    wb.save(_wb_dir("flecon") / "flecon_edge.xlsx")

    # EXTRACTOR registry (source_label drives signature matching). REORDERED vs
    # columns (F=ALL BLACK registered AFTER G=SLING) to prove signature mapping
    # survives a reshuffle a positional map would break. F/G share row-5 "FG";
    # the combined row3+row5 signature ("ALL BLACK FG" / "BLACK SLING 6X50 FG")
    # disambiguates via the contains-fallback pass.
    bag_type_registry = [
        {"code": "KURARAY_590", "source_label": "590 kls(Kuraray)", "source_column": "C",
         "sort_order": 1, "label": "Kuraray 590"},
        {"code": "UNUSABLE", "source_label": "Un-usable bag", "source_column": "D",
         "sort_order": 2, "label": "Un-usable bag"},
        {"code": "FUTAMURA_550", "source_label": "550 kls(FUTAMURA)", "source_column": "E",
         "sort_order": 3, "label": "Futamura 550"},
        {"code": "FG_BLACK_SLING_6X50", "source_label": "BLACK SLING 6X50", "source_column": "G",
         "sort_order": 5, "label": "FG Black Sling"},
        {"code": "FG_ALL_BLACK", "source_label": "ALL BLACK", "source_column": "F",
         "sort_order": 4, "label": "FG All Black"},
    ]
    # CLASSIFIER registry: {id, code}.
    bag_types = [
        {"id": "id-kuraray", "code": "KURARAY_590"},
        {"id": "id-unusable", "code": "UNUSABLE"},
        {"id": "id-futamura", "code": "FUTAMURA_550"},
        {"id": "id-sling", "code": "FG_BLACK_SLING_6X50"},
        {"id": "id-allblack", "code": "FG_ALL_BLACK"},
    ]
    # DB movements: make 2026-06-10 a DATE_CHANGED (qty off by 1 vs sheet) and add a
    # NOOP day (2026-05-05) whose multiset exactly matches the sheet's... but the
    # sheet has no 2026-05-05, so instead we prove NOOP via reorder within a day
    # in a SEPARATE case. Here: 2026-06-10 in DB differs by 1 -> DATE_CHANGED.
    movements = [
        # DB has KURARAY_590 -1 on 06-10 (sheet says -2) -> multiset differs -> DATE_CHANGED
        {"id": "m1", "transaction_date": "2026-06-10", "particular": "BAGGED POWDER",
         "bag_type_id": "id-kuraray", "qty_delta": -1},
        {"id": "m2", "transaction_date": "2026-06-10", "particular": "RS 1 ZAMBOANGA",
         "bag_type_id": "id-unusable", "qty_delta": -4},
        {"id": "m3", "transaction_date": "2026-06-10", "particular": "BLEND RECOUNT",
         "bag_type_id": "id-kuraray", "qty_delta": 5},
        {"id": "m4", "transaction_date": "2026-06-10", "particular": "BLEND RECOUNT",
         "bag_type_id": "id-futamura", "qty_delta": 7},
        # 2026-07-01 absent in DB -> NEW day
    ]
    _write_dbw("flecon", "flecon_edge", {
        "movements": movements, "bag_types": bag_types,
        "bag_type_registry": bag_type_registry, "view_balance": [],
    })

    # ── NOOP-by-reorder case: identical multiset, different row order in DB ──
    wb2 = openpyxl.Workbook()
    w2 = wb2.active
    w2.title = "JANUARY 2026"
    w2["E1"] = " FLECON BAG MOVEMENT"
    w2["A4"] = "DATE"; w2["B4"] = "PARTICULAR"; w2["C4"] = "FLECON BAG"
    w2["C5"] = "590 kls(Kuraray)"
    w2["D5"] = "Un-usable bag"
    r = 8
    w2.cell(r, 1, "2026-06-20"); w2.cell(r, 2, "A"); w2.cell(r, 3, -2); r += 1
    w2.cell(r, 2, "B"); w2.cell(r, 3, -2); r += 1  # duplicate identical (particular differs)
    w2.cell(r, 2, "C"); w2.cell(r, 4, -5); r += 1
    wb2.save(_wb_dir("flecon") / "flecon_noop_reorder.xlsx")
    # DB: SAME multiset, DIFFERENT emission order -> must be DUPLICATE_NOOP
    movements2 = [
        {"id": "n3", "transaction_date": "2026-06-20", "particular": "C",
         "bag_type_id": "id-unusable", "qty_delta": -5},
        {"id": "n1", "transaction_date": "2026-06-20", "particular": "A",
         "bag_type_id": "id-kuraray", "qty_delta": -2},
        {"id": "n2", "transaction_date": "2026-06-20", "particular": "B",
         "bag_type_id": "id-kuraray", "qty_delta": -2},
    ]
    _write_dbw("flecon", "flecon_noop_reorder", {
        "movements": movements2,
        "bag_types": [
            {"id": "id-kuraray", "code": "KURARAY_590"},
            {"id": "id-unusable", "code": "UNUSABLE"},
        ],
        "bag_type_registry": [
            {"code": "KURARAY_590", "source_label": "590 kls(Kuraray)", "source_column": "C",
             "sort_order": 1, "label": "Kuraray 590"},
            {"code": "UNUSABLE", "source_label": "Un-usable bag", "source_column": "D",
             "sort_order": 2, "label": "Un-usable bag"},
        ],
        "view_balance": [],
    })
    print("  ✓ flecon: flecon_edge.xlsx, flecon_noop_reorder.xlsx")


def _deliveries_sheet(ws) -> None:
    """Write the fixed deliveries header (rows 1-5) so find_header_row locates it
    (needs cols 1-6 concat to contain DATE OF/DELIVERY + SUPPLIER/SAMPLE)."""
    ws["A1"] = "JULY  2026 DELIVERIES"
    hdr2 = ["DATE", "DATE OF", "Sample", "Block", "block", "TRUCK", "Delivery",
            "Delivery", "Moisture", "Grit", "    BULK DENSITY", None, "VOLATILE",
            "ASH", "FIXED", None]
    hdr3 = ["ANALYZED", "DELIVERY", "Supplier name", None, "locator", "PLATE",
            "(kilo)", "(sack)", "CONTENT", None, "ASTM", "JIS", "MATTER",
            "CONTENT", "CARBON", None]
    for c, v in enumerate(hdr2, 1):
        ws.cell(2, c, v)
    for c, v in enumerate(hdr3, 1):
        ws.cell(3, c, v)


def _deliveries_row(ws, r, *, date, supplier, block, block_loc, plate, kg, sacks,
                    mc=11.0, grit=3.0, bd_astm=0.58, bd_jis=0.60, vm=13.0, ash=3.0,
                    fc=82.0, remarks=None):
    ws.cell(r, 1, date); ws.cell(r, 2, date); ws.cell(r, 3, supplier)
    ws.cell(r, 4, block); ws.cell(r, 5, block_loc); ws.cell(r, 6, plate)
    ws.cell(r, 7, kg); ws.cell(r, 8, sacks)
    ws.cell(r, 9, mc); ws.cell(r, 10, grit); ws.cell(r, 11, bd_astm)
    ws.cell(r, 12, bd_jis); ws.cell(r, 13, vm); ws.cell(r, 14, ash); ws.cell(r, 15, fc)
    ws.cell(r, 16, remarks)


def build_deliveries() -> None:
    """L-033 A-19C replay + date-carry-forward gap + off-format block_loc.

    Scenario:
      - Row: 2026-07-04 delivery, block A-19C, plate 'MAV 9202', 20640kg, remark
        'PILED IN JUNE BLOCK 9'. The DB already holds JUNE-26-BLK9 as a batch,
        AND a deliveries row same (date, truck, weight) at A-19C under
        JUNE-26-BLK9 -> L-033a demotes to dup_noops (phantom month-boundary name).
      - Row: 2026-07-04 second truck at A-19C, remark 'PILED IN JUNE BLOCK 9',
        DIFFERENT weight -> no dtw match -> L-033b remark re-maps its
        extractor-derived JULY code to the existing JUNE-26-BLK9, then inserts.
      - Row with a blank date (carry-forward from the row above).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JULY 26"
    _deliveries_sheet(ws)
    from datetime import datetime
    d = datetime(2026, 7, 4)
    # Row 6: the A-19C phantom (L-033a dup_noop) — matches a DB row by (date,truck,weight) at same loc.
    _deliveries_row(ws, 6, date=d, supplier="Ornales", block="B09", block_loc="A-19C",
                    plate="MAV 9202", kg=20640, sacks=580, remarks="PILED IN JUNE BLOCK 9")
    # Row 7: blank date -> carry-forward; different truck+weight -> L-033b remap path.
    _deliveries_row(ws, 7, date=None, supplier="Ornales", block="B09", block_loc="A-19C",
                    plate="CBN 2192", kg=10870, sacks=270, remarks="PILED IN JUNE BLOCK 9")
    # Row 8: a clean unrelated NEW row on a fresh date (no remark, off nothing).
    _deliveries_row(ws, 8, date=datetime(2026, 7, 5), supplier="Tagat", block="B02",
                    block_loc="D-19B", plate="LFF 835", kg=12725, sacks=286)
    _wb_dir("deliveries")
    wb.save(_wb_dir("deliveries") / "deliveries_l033.xlsx")

    # DB window: the existing JUNE batch + the A-19C row already recorded under JUNE-26-BLK9.
    db_deliveries = [{
        "id": "db-a19c", "transaction_date": "2026-07-04", "supplier": "Ornales",
        "batch_code": "JUNE-26-BLK9", "block_loc": "A-19C", "truck_plate": "MAV 9202",
        "sacks": 580, "weight_kg": 20640.0, "cost_basis": 38.0, "remarks": None,
        "lab_results": {"mc": 11.0, "grit": 3.0, "bd_astm": 0.58, "bd_jis": 0.60,
                        "vm": 13.0, "ash": 3.0, "fc": 82.0},
    }]
    batch_codes = ["JUNE-26-BLK9", "JULY-26-BLK9"]
    _write_dbw("deliveries", "deliveries_l033", {
        "deliveries": db_deliveries, "batch_codes": batch_codes,
    })
    print("  ✓ deliveries: deliveries_l033.xlsx")


def build_gsheet() -> None:
    """gsheet materiality + null<->0 + out-of-scope + destination typo.

    RC IN tab (header row 7, data 8+). RC OUT tab (header row 4, data 5+).
    Curated DB window makes:
      - one RC IN row an IMMATERIAL change (sacks null<->0) -> NOOP with note,
      - one RC IN row a MATERIAL change (remarks differ) -> CHANGED (sheet-wins),
      - one RC IN row OUT-OF-SCOPE (2024) -> dropped to out_of_scope,
      - one RC OUT row a destination typo 'MAN' -> normalized to MAIN.
    """
    from datetime import datetime
    wb = openpyxl.Workbook()
    rc_in = wb.active
    rc_in.title = "RC IN"
    rc_in_hdr = ["STATE", "WHSE", "DATE", "SUPPLIER", "BLOCK", "BLOCK LOC", "TRK",
                 "WT", "SKS", "MC", "GRIT", "ASTM", "JIS", "VM", "ASH", "FC", "REMARKS"]
    for c, v in enumerate(rc_in_hdr, 1):
        rc_in.cell(7, c, v)

    def rin(r, *, date, supplier, batch, loc, trk, wt, sks=None, mc=None, remarks=None):
        rc_in.cell(r, 3, date); rc_in.cell(r, 4, supplier); rc_in.cell(r, 5, batch)
        rc_in.cell(r, 6, loc); rc_in.cell(r, 7, trk); rc_in.cell(r, 8, wt)
        rc_in.cell(r, 9, sks); rc_in.cell(r, 10, mc); rc_in.cell(r, 17, remarks)

    # r8: immaterial sacks null<->0 (sheet sks blank, DB sacks=0) -> NOOP w/ note
    rin(8, date=datetime(2026, 6, 10), supplier="ACME", batch="JUNE-26-BLK1",
        loc="A-1A", trk="ABC 111", wt=10000.0, sks=None, mc=11.5)
    # r9: material remarks change (sheet 'FIXED' vs DB null) -> CHANGED
    rin(9, date=datetime(2026, 6, 11), supplier="ACME", batch="JUNE-26-BLK2",
        loc="A-2A", trk="ABC 222", wt=20000.0, sks=100, mc=12.0, remarks="CORRECTED")
    # r10: OUT OF SCOPE (2024) -> dropped silently to out_of_scope
    rin(10, date=datetime(2024, 12, 31), supplier="OLD", batch="DEC-24-BLK9",
        loc="C-9A", trk="OLD 999", wt=5000.0)
    # r11: a genuinely NEW row (absent from DB)
    rin(11, date=datetime(2026, 6, 12), supplier="NEWCO", batch="JUNE-26-BLK7",
        loc="A-7A", trk="NEW 777", wt=15000.0, sks=300, mc=10.0)

    rc_out = wb.create_sheet("RC OUT")
    rc_out_hdr = ["DATE", "BATCH", "BLOCK", "WT", "PLANT/ETC", "REMARKS", "BLOCK LOC",
                  None, "MC", "MC WTD", "DAY"]
    for c, v in enumerate(rc_out_hdr, 1):
        if v is not None:
            rc_out.cell(4, c, v)

    def rout(r, *, date, prod_batch, batch_code, wt, dest, remarks=None):
        rc_out.cell(r, 1, date); rc_out.cell(r, 2, prod_batch); rc_out.cell(r, 3, batch_code)
        rc_out.cell(r, 4, wt); rc_out.cell(r, 5, dest); rc_out.cell(r, 6, remarks)

    # r5: destination typo 'MAN' -> normalized MAIN; matches a DB rc_out row -> NOOP
    rout(5, date=datetime(2026, 6, 10), prod_batch="JUNE", batch_code="JUNE-26-BLK1",
         wt=3000.0, dest="MAN")
    # r6: a new rc_out consumption
    rout(6, date=datetime(2026, 6, 12), prod_batch="JUNE", batch_code="JUNE-26-BLK7",
         wt=4000.0, dest="MAIN")

    _wb_dir("gsheet")
    wb.save(_wb_dir("gsheet") / "gsheet_edge.xlsx")

    # DB window (RC IN resolves batch against deliveries.batch_code in-window;
    # RC OUT resolves batch_id against a batches lookup).
    db_deliveries = [
        # matches r8 exactly except sacks=0 (sheet blank) -> immaterial NOOP
        {"id": "g1", "transaction_date": "2026-06-10", "supplier": "ACME",
         "batch_code": "JUNE-26-BLK1", "block_loc": "A-1A", "truck_plate": "ABC 111",
         "sacks": 0, "weight_kg": 10000.0, "cost_basis": 30.0, "remarks": None,
         "lab_results": {"mc": 11.5}},
        # matches r9 except remarks (DB null vs sheet 'CORRECTED') -> material CHANGED
        {"id": "g2", "transaction_date": "2026-06-11", "supplier": "ACME",
         "batch_code": "JUNE-26-BLK2", "block_loc": "A-2A", "truck_plate": "ABC 222",
         "sacks": 100, "weight_kg": 20000.0, "cost_basis": 31.0, "remarks": None,
         "lab_results": {"mc": 12.0}},
    ]
    db_rc_out = [
        {"id": "o1", "transaction_date": "2026-06-10", "batch_id": "bid-blk1",
         "production_batch": "JUNE", "destination": "MAIN", "weight_kg": 3000.0,
         "block_loc": "A-1A", "remarks": None},
    ]
    batch_lookup = {"JUNE-26-BLK1": "bid-blk1", "JUNE-26-BLK7": "bid-blk7"}
    _write_dbw("gsheet", "gsheet_edge", {
        "deliveries": db_deliveries, "rc_out": db_rc_out, "batch_lookup": batch_lookup,
    })
    print("  ✓ gsheet: gsheet_edge.xlsx")


def _proposed_section(ws, R, *, whse, block_date, block_no, gross, pallet, net,
                      day_total, status=None, remarks=None):
    """Write one 7-row PROPOSED block section starting at row R (mirrors the real
    geometry: col A labels, col B pallet values, col K/L/M stats)."""
    ws.cell(R + 0, 1, "WHSE #"); ws.cell(R + 0, 2, whse)
    ws.cell(R + 0, 11, "STRT. BAL"); ws.cell(R + 0, 12, 9999)
    if status is not None:
        ws.cell(R + 0, 13, status)
    ws.cell(R + 1, 1, "BLOCK DATE"); ws.cell(R + 1, 2, block_date)
    ws.cell(R + 1, 11, "DAY TOTAL"); ws.cell(R + 1, 12, day_total)
    ws.cell(R + 2, 1, "BLOCK NO."); ws.cell(R + 2, 2, block_no)
    ws.cell(R + 2, 11, "END BAL."); ws.cell(R + 2, 12, 0)
    ws.cell(R + 3, 1, "Gross weight"); ws.cell(R + 3, 2, gross)
    ws.cell(R + 3, 11, "REMARKS")
    if remarks is not None:
        ws.cell(R + 3, 12, remarks)
    ws.cell(R + 4, 1, "Pallet"); ws.cell(R + 4, 2, pallet)
    ws.cell(R + 5, 1, "Net"); ws.cell(R + 5, 2, net)
    # R+6 blank separator


def build_rc_out() -> None:
    """PROPOSED DAILY REPORT: an UNMAPPED batch section + a sub-watermark FLAGGED
    row + a clean resolvable section.

    One sheet 'JULY 4' (2026). Sections:
      - block_no 9, block_date 2026-07-01 -> JULY-26-BLK9 (exists) -> NEW.
      - block_no 99 -> JULY-26-BLK99 (NOT in batch_lookup) -> UNMAPPED.
    A watermark of 2026-07-10 makes the 2026-07-04 rows sub-watermark: any row
    with no natural-key match is FLAGGED (L-019), never NEW.
    """
    from datetime import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JULY 4"
    ws["A1"] = "DAILY REPORT OF RAW CHARCOAL FED "
    ws["I1"] = "DATE "; ws["J1"] = "JULY 4,2026"
    _proposed_section(ws, 4, whse="C-6A", block_date=datetime(2026, 7, 1),
                      block_no="# 9", gross=1498, pallet=84, net=1414, day_total=1414)
    _proposed_section(ws, 12, whse="A-5B", block_date=datetime(2026, 7, 1),
                      block_no="# 99", gross=1203, pallet=70, net=1203, day_total=1203)
    _wb_dir("rc_out")
    wb.save(_wb_dir("rc_out") / "rc_out_edge.xlsx")

    # batch_lookup has BLK9 but NOT BLK99 -> BLK99 section UNMAPPED.
    batch_lookup = {"JULY-26-BLK9": "bid-blk9"}
    # rc_out DB rows empty in-window (so both sections are candidate NEW), but the
    # watermark makes the 2026-07-04 rows sub-watermark -> FLAGGED (L-019).
    _write_dbw("rc_out", "rc_out_edge", {
        "rc_out": [], "batch_lookup": batch_lookup, "rc_out_sums": {},
    })
    print("  ✓ rc_out: rc_out_edge.xlsx")


def build_production() -> None:
    """MC Daily Production Report: L-026 duplicate-run combine + a >=60-min
    downtime day (deviation #5, the DB CHECK-violating scenario) + a dropped
    grade + a blank-shift run (L-025 default).

    Coordinates (extract_daily_production.py): runs rows 8-12 (D grade, E sacks,
    G ttl, H shift); TOTAL row 13 (C13='TOTAL', G13 day total); downtime C24
    category, C27 ranges, E27 minutes, F27 reasons.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "07-04-26"  # MM-DD-YY -> 2026-07-04
    # runs header (row 7 not parsed for data)
    ws["C7"] = "SHIFT"; ws["D7"] = "MESH SIZES"; ws["E7"] = "#SACKS/BAGS"
    ws["F7"] = "#KILOS"; ws["G7"] = "TOTAL"; ws["H7"] = "SHIFT"
    # r8/r9: SAME customer(CEBU)+grade(3X50)+shift resolution -> L-026 combine.
    ws.cell(8, 4, "CEBU 3X50"); ws.cell(8, 5, 630); ws.cell(8, 7, 16380); ws.cell(8, 8, "DAY SHIFT")
    ws.cell(9, 4, "CEBU 3X50"); ws.cell(9, 5, 218); ws.cell(9, 7, 5668); ws.cell(9, 8, "DAY SHIFT")
    # r10: dropped grade (KOREA POWDER not in VALID_GRADES) -> dropped_grades
    ws.cell(10, 4, "KOREA POWDER (BAGGED)"); ws.cell(10, 5, 100); ws.cell(10, 7, 5000); ws.cell(10, 8, "DAY SHIFT")
    # r11: blank-shift run (L-025 default to Morning) with a valid grade
    ws.cell(11, 4, "KURARAY 6X50"); ws.cell(11, 5, 200); ws.cell(11, 7, 10000)  # shift H blank
    # TOTAL row 13
    ws.cell(13, 3, "TOTAL"); ws.cell(13, 7, 37048)
    # downtime: category C24, ranges C27 (multiline), minutes E27 (parallel), reasons F27
    ws.cell(24, 3, "REPAIR")
    ws.cell(27, 3, "8:00-9:00\n10:00-10:30")  # two ranges
    ws.cell(27, 5, "60 MINUTES\n65 MINUTES")  # sums to 125 -> >=60 -> deviation #5
    ws.cell(27, 6, "belt change\nmotor")
    _wb_dir("production")
    wb.save(_wb_dir("production") / "production_mc_edge.xlsx")

    # DB window: empty child tables + no shifts -> all NEW (shift upsert path).
    _write_dbw("production", "production_downtime_ge60", {
        "shifts": [], "runs": [], "downtime": [], "waste": [],
        "electricity": [], "trucks": [],
    })
    print("  ✓ production: production_mc_edge.xlsx")


BUILDERS = {
    "flecon": build_flecon,
    "deliveries": build_deliveries,
    "gsheet": build_gsheet,
    "rc_out": build_rc_out,
    "production": build_production,
}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--type")
    args = ap.parse_args()
    types = [args.type] if args.type else sorted(BUILDERS.keys())
    for rt in types:
        if rt not in BUILDERS:
            print(f"no synthetic builder for {rt} (yet)")
            continue
        BUILDERS[rt]()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
